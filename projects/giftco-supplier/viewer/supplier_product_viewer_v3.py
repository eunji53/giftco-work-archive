# -*- coding: utf-8 -*-
"""
공급사별 상품 카테고리/상품 조회 툴 - 2개 파일 기준 버전

필수 파일
- 공급사 목록/공급사 정보: supplier_detail_result.xlsx
- 상품 조회: products_with_supplier_info.xlsx

실행 전 설치
    pip install pandas openpyxl

실행
    python supplier_product_viewer_v3.py

exe 생성
    pyinstaller --onefile --windowed --name supplier_product_viewer supplier_product_viewer_v3.py
"""

from __future__ import annotations

import re
import json
import sys
import webbrowser
import traceback
from pathlib import Path
from datetime import datetime
import tkinter as tk
from tkinter import ttk, filedialog, messagebox
from tkinter.scrolledtext import ScrolledText

try:
    import pandas as pd
except ImportError:
    pd = None

try:
    from openpyxl import load_workbook
except ImportError:
    load_workbook = None

def get_app_dir() -> Path:
    """py 파일 실행/EXE 실행 모두에서 프로그램 폴더를 반환합니다."""
    if getattr(sys, "frozen", False):
        return Path(sys.executable).resolve().parent
    return Path(__file__).resolve().parent


# =========================
# 1. 파일 설정
# =========================
# 상품 파일은 프로그램 폴더 기준 ../data/ 에 고정으로 둡니다 (crawler와 공유하는 데이터 폴더).
DATA_DIR = "../data"
JOIN_FILE = f"{DATA_DIR}/products_with_supplier_info.xlsx"

# 공급사 파일은 실행 시 사용자가 직접 선택합니다. 마지막 선택 경로는 config에 저장됩니다.
CONFIG_FILE = f"{DATA_DIR}/supplier_product_viewer_config.json"

# 연락상태는 공급사 파일을 직접 수정하지 않고 별도 파일에 저장합니다.
CONTACT_STATUS_FILE = "공급사_연락상태.xlsx"

CONTACT_STATUS_COLUMNS = [
    "업체코드",
    "업체명",
    "담당자",
    "연락여부",
    "연락일자",
    "연락수단",
    "연락결과",
    "재연락예정일",
    "메모",
    "최종수정일시",
]

CONTACT_STATUS_OPTIONS = ["전체", "미연락", "연락완료", "재연락필요", "연락불가"]

# 미분류 상품을 카테고리/대분류 요약에 표시할지 여부
# False로 바꾸면 카테고리가 빈 상품은 요약/카테고리 목록에서 숨깁니다.
SHOW_UNCATEGORIZED = True

# 상품 리스트의 진열상태 필터 옵션
# 엑셀의 "진열" 값이 비어 있으면 화면에서는 "미분류"로 표시/필터링합니다.
DISPLAY_STATUS_OPTIONS = ["전체", "진열", "단종", "중지", "품절", "미분류"]

# 왼쪽 공급사 목록에 표시할 컬럼
SUPPLIER_LIST_COLUMNS = ["연락", "업체코드", "업체명", "상품수", "연락여부", "연락일자"]

# 오른쪽 공급사 정보 영역에 표시할 항목
SUPPLIER_LEFT_FIELDS = [
    "공급처_사업자번호",
    "공급처_대표자",
    "공급처_업태",
    "공급처_종목",
    "공급처_주소",
]

SUPPLIER_RIGHT_FIELDS = [
    "전화번호",
    "휴대폰번호",
    "팩스",
    "이메일",
]

FIELD_LABELS = {
    "업체명": "업체명",
    "공급처_사업자번호": "사업자번호",
    "공급처_대표자": "대표자",
    "공급처_업태": "업태",
    "공급처_종목": "종목",
    "공급처_주소": "주소",
    "상품수": "상품수",
    "카테고리수": "카테고리수",
    "전화번호": "전화번호",
    "휴대폰번호": "휴대폰번호",
    "팩스": "팩스",
    "이메일": "이메일",
}

# 상품 리스트에 표시할 컬럼
PRODUCT_COLUMNS = [
    "상품코드", "진열", "상품명", "카테고리", "공급가", "판매가1", "판매가7",
    "마진율", "노출", "최초등록일", "최근수정일", "상품링크"
]


# =========================
# 2. 공통 유틸
# =========================
def clean_text(value) -> str:
    if value is None:
        return ""
    try:
        if pd is not None and pd.isna(value):
            return ""
    except Exception:
        pass
    text = str(value).strip()
    if text.lower() in {"nan", "none", "nat"}:
        return ""
    return text


def norm_key(value) -> str:
    text = clean_text(value)
    text = re.sub(r"\s+", "", text)
    return text.upper()


def safe_int(value, default=0) -> int:
    try:
        text = clean_text(value).replace(",", "")
        if text == "":
            return default
        return int(float(text))
    except Exception:
        return default


def normalize_df(df: "pd.DataFrame") -> "pd.DataFrame":
    df = df.copy()
    df.columns = [clean_text(c) for c in df.columns]
    df = df.fillna("")
    for col in df.columns:
        if df[col].dtype == object:
            df[col] = df[col].astype(str).str.strip()
    return df


def read_excel_fast(path: Path) -> "pd.DataFrame":
    if load_workbook is None:
        raise ImportError("openpyxl이 설치되어 있지 않습니다. pip install openpyxl 후 다시 실행하세요.")
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)
    try:
        headers = next(rows_iter)
    except StopIteration:
        return pd.DataFrame()
    headers = [clean_text(h) for h in headers]
    rows = list(rows_iter)
    return normalize_df(pd.DataFrame(rows, columns=headers))


def category_display(value: str) -> str:
    cat = clean_text(value)
    return cat if cat else "미분류"


def major_display(value: str) -> str:
    cat = clean_text(value)
    if not cat:
        return "미분류"
    return cat.split(">", 1)[0].strip() or "미분류"


def display_status(value: str) -> str:
    status = clean_text(value)
    return status if status else "미분류"


# =========================
# 3. 메인 앱
# =========================
class SupplierProductViewer(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("공급사별 상품 카테고리 조회 툴")
        self.geometry("1850x1000")
        self.minsize(1300, 800)

        if pd is None:
            messagebox.showerror(
                "pandas 설치 필요",
                "이 프로그램은 pandas/openpyxl이 필요합니다.\n\n명령 프롬프트에서 아래 명령을 실행하세요.\n\npip install pandas openpyxl"
            )
            self.destroy()
            return

        self.app_dir = get_app_dir()
        self.data_dir = self.app_dir  # 기존 코드 호환용
        self.join_path = self.app_dir / JOIN_FILE
        self.config_path = self.app_dir / CONFIG_FILE
        self.contact_status_path = self.app_dir / CONTACT_STATUS_FILE
        self.supplier_path: Path | None = None

        self.supplier_df = pd.DataFrame()
        self.goods_df = pd.DataFrame()
        self.contact_status_df = pd.DataFrame()
        self.assignment_df = pd.DataFrame()  # 기존 코드 호환용
        self.supplier_list_df = pd.DataFrame()

        self.load_config()

        self.selected_supplier_code = ""
        self.selected_supplier_name = ""
        self.current_product_df = pd.DataFrame(columns=PRODUCT_COLUMNS)
        self.current_product_filtered_df = pd.DataFrame(columns=PRODUCT_COLUMNS)
        self.current_category_filter = ""
        self.current_major_filter = ""
        self.current_display_status_filter = "전체"
        self.product_display_limit = 1000
        self.product_display_count = 1000

        self._setup_style()
        self._build_layout()
        self.after(250, self.force_layout)
        self.after(400, self.load_data)

    # ---------------- UI ----------------
    def _setup_style(self):
        style = ttk.Style(self)
        try:
            style.theme_use("clam")
        except Exception:
            pass
        base_font = ("맑은 고딕", 9)
        style.configure("Treeview", rowheight=24, font=base_font)
        style.configure("Treeview.Heading", font=("맑은 고딕", 9, "bold"))
        style.configure("TLabel", font=base_font)
        style.configure("TButton", font=base_font)
        style.configure("TLabelframe.Label", font=("맑은 고딕", 10, "bold"))

    def _build_layout(self):
        top_bar = ttk.Frame(self, padding=(8, 6))
        top_bar.pack(fill=tk.X)
        ttk.Label(top_bar, text="공급사 파일:").pack(side=tk.LEFT)
        self.path_label = ttk.Label(top_bar, text=self._supplier_path_text(), foreground="#555")
        self.path_label.pack(side=tk.LEFT, padx=(6, 12), fill=tk.X, expand=True)
        ttk.Button(top_bar, text="공급사 파일 선택/재로드", command=self.choose_supplier_file_and_reload).pack(side=tk.RIGHT, padx=(6, 0))
        ttk.Button(top_bar, text="전체 새로고침", command=self.load_data).pack(side=tk.RIGHT)

        self.status_var = tk.StringVar(value="준비 중")
        ttk.Label(self, textvariable=self.status_var, anchor="w", padding=(8, 4)).pack(fill=tk.X)

        self.main_paned = ttk.PanedWindow(self, orient=tk.HORIZONTAL)
        self.main_paned.pack(fill=tk.BOTH, expand=True, padx=8, pady=8)

        # 왼쪽: 공급사 선택용. 일부러 좁게 사용.
        left = ttk.Frame(self.main_paned)
        self.main_paned.add(left, weight=1)

        supplier_box = ttk.LabelFrame(left, text="공급사 목록", padding=6)
        supplier_box.pack(fill=tk.BOTH, expand=True)

        search_frame = ttk.Frame(supplier_box)
        search_frame.pack(fill=tk.X, pady=(0, 5))
        ttk.Label(search_frame, text="공급사 검색").pack(side=tk.LEFT)
        self.supplier_search_var = tk.StringVar()
        self.supplier_search_var.trace_add("write", lambda *args: self.refresh_supplier_list())
        ttk.Entry(search_frame, textvariable=self.supplier_search_var).pack(side=tk.LEFT, fill=tk.X, expand=True, padx=5)

        ttk.Label(search_frame, text="연락상태").pack(side=tk.LEFT, padx=(6, 4))
        self.contact_filter_var = tk.StringVar(value="전체")
        self.contact_filter_combo = ttk.Combobox(
            search_frame,
            textvariable=self.contact_filter_var,
            values=CONTACT_STATUS_OPTIONS,
            state="readonly",
            width=10,
        )
        self.contact_filter_combo.pack(side=tk.LEFT)
        self.contact_filter_combo.bind("<<ComboboxSelected>>", lambda event: self.refresh_supplier_list())

        # 담당자/상태 필터 UI는 제거. 변수만 유지해서 기존 필터 로직이 깨지지 않게 처리.
        self.owner_filter_var = tk.StringVar(value="전체")
        self.status_filter_var = tk.StringVar(value="전체")

        button_frame = ttk.Frame(supplier_box)
        button_frame.pack(fill=tk.X, pady=(0, 5))
        ttk.Button(button_frame, text="연락상태 목록 저장", command=self.export_assignment_template).pack(side=tk.RIGHT)

        supplier_table_frame = ttk.Frame(supplier_box)
        supplier_table_frame.pack(fill=tk.BOTH, expand=True)
        self.supplier_tree = ttk.Treeview(
            supplier_table_frame,
            columns=tuple(SUPPLIER_LIST_COLUMNS),
            show="headings",
            selectmode="browse",
        )
        self._config_tree(self.supplier_tree, {
            "연락": 45,
            "업체코드": 70,
            "업체명": 165,
            "상품수": 60,
            "연락여부": 90,
            "연락일자": 95,
        })
        self.supplier_tree.tag_configure("contact_none", background="#FFFFFF")
        self.supplier_tree.tag_configure("contact_done", background="#E2F0D9")
        self.supplier_tree.tag_configure("contact_retry", background="#FFF2CC")
        self.supplier_tree.tag_configure("contact_fail", background="#E7E6E6")
        sy = ttk.Scrollbar(supplier_table_frame, orient=tk.VERTICAL, command=self.supplier_tree.yview)
        sx = ttk.Scrollbar(supplier_table_frame, orient=tk.HORIZONTAL, command=self.supplier_tree.xview)
        self.supplier_tree.configure(yscrollcommand=sy.set, xscrollcommand=sx.set)
        self.supplier_tree.grid(row=0, column=0, sticky="nsew")
        sy.grid(row=0, column=1, sticky="ns")
        sx.grid(row=1, column=0, sticky="ew")
        supplier_table_frame.rowconfigure(0, weight=1)
        supplier_table_frame.columnconfigure(0, weight=1)
        self.supplier_tree.bind("<<TreeviewSelect>>", self.on_supplier_select)

        # 오른쪽: 상세/요약/상품
        self.right_paned = ttk.PanedWindow(self.main_paned, orient=tk.VERTICAL)
        self.main_paned.add(self.right_paned, weight=9)

        self.top_right_paned = ttk.PanedWindow(self.right_paned, orient=tk.HORIZONTAL)
        self.right_paned.add(self.top_right_paned, weight=3)

        info_box = ttk.LabelFrame(self.top_right_paned, text="공급사 정보", padding=6)
        self.top_right_paned.add(info_box, weight=5)

        contact_button_frame = ttk.Frame(info_box)
        contact_button_frame.pack(fill=tk.X, pady=(0, 6))
        ttk.Button(contact_button_frame, text="연락완료", command=lambda: self.update_contact_status("연락완료")).pack(side=tk.LEFT, padx=(0, 4))
        ttk.Button(contact_button_frame, text="재연락 필요", command=lambda: self.update_contact_status("재연락필요")).pack(side=tk.LEFT, padx=(0, 4))
        ttk.Button(contact_button_frame, text="연락불가", command=lambda: self.update_contact_status("연락불가")).pack(side=tk.LEFT, padx=(0, 4))
        ttk.Button(contact_button_frame, text="미연락", command=lambda: self.update_contact_status("미연락")).pack(side=tk.LEFT, padx=(0, 4))

        # self.info_text = ScrolledText(info_box, height=12, font=("맑은 고딕", 9), wrap=tk.WORD)
        # self.info_text.pack(fill=tk.BOTH, expand=True)
        # self.info_text.configure(state=tk.DISABLED, spacing1=2, spacing3=4, padx=8, pady=6)
        self.info_text = ScrolledText(
            info_box,
            height=11,
            font=("맑은 고딕", 10),
            wrap=tk.WORD
        )
        self.info_text.pack(fill=tk.BOTH, expand=True)

        self.info_text.configure(
            foreground="#111111",
            background="#FFFFFF",
            spacing1=2,
            spacing3=5,
            padx=10,
            pady=8
        )

        self.info_text.configure(state=tk.DISABLED)

        major_box = ttk.LabelFrame(self.top_right_paned, text="대분류 요약", padding=6)
        self.top_right_paned.add(major_box, weight=3)
        self.major_tree = ttk.Treeview(major_box, columns=("대분류", "상품수"), show="headings", selectmode="browse")
        self._config_tree(self.major_tree, {"대분류": 320, "상품수": 90})
        my = ttk.Scrollbar(major_box, orient=tk.VERTICAL, command=self.major_tree.yview)
        self.major_tree.configure(yscrollcommand=my.set)
        self.major_tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        my.pack(side=tk.RIGHT, fill=tk.Y)
        self.major_tree.bind("<<TreeviewSelect>>", self.on_major_select)

        self.bottom_paned = ttk.PanedWindow(self.right_paned, orient=tk.HORIZONTAL)
        self.right_paned.add(self.bottom_paned, weight=6)

        category_box = ttk.LabelFrame(self.bottom_paned, text="카테고리 목록", padding=6)
        self.bottom_paned.add(category_box, weight=3)
        cat_top = ttk.Frame(category_box)
        cat_top.pack(fill=tk.X, pady=(0, 5))
        ttk.Button(cat_top, text="카테고리/대분류 필터 해제", command=self.clear_category_filter).pack(side=tk.LEFT)
        self.category_filter_label = ttk.Label(cat_top, text="필터 없음", foreground="#555")
        self.category_filter_label.pack(side=tk.LEFT, padx=8)

        cat_frame = ttk.Frame(category_box)
        cat_frame.pack(fill=tk.BOTH, expand=True)
        self.category_tree = ttk.Treeview(
            cat_frame,
            # columns=("카테고리", "개수", "대", "중", "소", "세"),
            columns=("카테고리", "상품수"),
            show="headings",
            selectmode="browse",
        )
        self._config_tree(self.category_tree, {
            "카테고리": 420,
            "상품수": 70
            # "개수": 70,
            # "대": 150,
            # "중": 150,
            # "소": 150,
            # "세": 120,
        })
        cy = ttk.Scrollbar(cat_frame, orient=tk.VERTICAL, command=self.category_tree.yview)
        cx = ttk.Scrollbar(cat_frame, orient=tk.HORIZONTAL, command=self.category_tree.xview)
        self.category_tree.configure(yscrollcommand=cy.set, xscrollcommand=cx.set)
        self.category_tree.grid(row=0, column=0, sticky="nsew")
        cy.grid(row=0, column=1, sticky="ns")
        cx.grid(row=1, column=0, sticky="ew")
        cat_frame.rowconfigure(0, weight=1)
        cat_frame.columnconfigure(0, weight=1)
        self.category_tree.bind("<<TreeviewSelect>>", self.on_category_select)

        product_box = ttk.LabelFrame(self.bottom_paned, text="상품 리스트", padding=6)
        self.bottom_paned.add(product_box, weight=5)
        product_top = ttk.Frame(product_box)
        product_top.pack(fill=tk.X, pady=(0, 5))
        ttk.Label(product_top, text="상품 검색").pack(side=tk.LEFT)
        self.product_search_var = tk.StringVar()
        self.product_search_var.trace_add("write", lambda *args: self.refresh_product_list())
        ttk.Entry(product_top, textvariable=self.product_search_var).pack(side=tk.LEFT, fill=tk.X, expand=True, padx=5)

        ttk.Label(product_top, text="진열상태").pack(side=tk.LEFT, padx=(8, 4))
        self.display_status_filter_var = tk.StringVar(value="전체")
        self.display_status_combo = ttk.Combobox(
            product_top,
            textvariable=self.display_status_filter_var,
            values=DISPLAY_STATUS_OPTIONS,
            state="readonly",
            width=8,
        )
        self.display_status_combo.pack(side=tk.LEFT, padx=(0, 5))
        self.display_status_combo.bind("<<ComboboxSelected>>", self.on_display_status_filter_change)

        ttk.Button(product_top, text="공급사 요약 저장", command=self.export_supplier_summary).pack(side=tk.RIGHT, padx=(5, 0))
        ttk.Button(product_top, text="상품목록 저장", command=self.export_current_products).pack(side=tk.RIGHT, padx=(5, 0))
        ttk.Button(product_top, text="링크 열기", command=self.open_selected_product_link).pack(side=tk.RIGHT, padx=(5, 0))
        ttk.Button(product_top, text="더 보기", command=self.show_more_products).pack(side=tk.RIGHT, padx=(5, 0))

        self.product_count_var = tk.StringVar(value="상품 0건")
        ttk.Label(product_box, textvariable=self.product_count_var, foreground="#555").pack(fill=tk.X, pady=(0, 4))

        product_frame = ttk.Frame(product_box)
        product_frame.pack(fill=tk.BOTH, expand=True)
        self.product_tree = ttk.Treeview(product_frame, columns=tuple(PRODUCT_COLUMNS), show="headings", selectmode="browse")
        self._config_tree(self.product_tree, {
            "상품코드": 90,
            "상품명": 420,
            "카테고리": 420,
            # "기본수량": 75,
            "공급가": 85,
            "판매가1": 85,
            "판매가7": 85,
            "마진율": 65,
            "진열": 60,
            "노출": 60,
            "최초등록일": 90,
            "최근수정일": 90,
            "상품링크": 260,
            # "이미지URL": 220,
        })
        py = ttk.Scrollbar(product_frame, orient=tk.VERTICAL, command=self.product_tree.yview)
        px = ttk.Scrollbar(product_frame, orient=tk.HORIZONTAL, command=self.product_tree.xview)
        self.product_tree.configure(yscrollcommand=py.set, xscrollcommand=px.set)
        self.product_tree.grid(row=0, column=0, sticky="nsew")
        py.grid(row=0, column=1, sticky="ns")
        px.grid(row=1, column=0, sticky="ew")
        product_frame.rowconfigure(0, weight=1)
        product_frame.columnconfigure(0, weight=1)
        self.product_tree.bind("<Double-1>", lambda event: self.open_selected_product_link())

    def force_layout(self):
        """초기 화면 비율을 강제로 조정합니다."""
        try:
            self.update_idletasks()
            w = max(self.winfo_width(), 1300)
            h = max(self.winfo_height(), 800)

            # 왼쪽 공급사 목록은 좁게
            left_width = min(300, max(230, int(w * 0.15)))
            self.main_paned.sashpos(0, left_width)

            # 공급사 정보/대분류 상단 영역 높이 확대
            self.right_paned.sashpos(0, max(320, int(h * 0.33)))

            # 공급사 정보 영역 넓게, 대분류 요약은 오른쪽
            self.top_right_paned.sashpos(0, max(680, int(w * 0.46)))

            # 카테고리 목록을 기존보다 조금 넓게
            self.bottom_paned.sashpos(0, max(400, int(w * 0.3)))
        except Exception:
            pass

    def _config_tree(self, tree: ttk.Treeview, widths: dict[str, int]):
        for col, width in widths.items():
            tree.heading(col, text=col, command=lambda c=col, t=tree: self.sort_treeview(t, c, False))

            if col in {"상품수", "연락", "연락여부", "연락일자"}:
                anchor = tk.CENTER
            elif col in {"카테고리수", "개수", "조인상품수", "공급가", "판매가1", "판매가7", "기본수량"}:
                anchor = tk.E
            else:
                anchor = tk.W

            tree.column(col, width=width, minwidth=40, anchor=anchor)

    # ---------------- Data loading ----------------
    def _supplier_path_text(self) -> str:
        if self.supplier_path and self.supplier_path.exists():
            return str(self.supplier_path)
        return "공급사 파일을 선택하세요. / 상품 파일은 프로그램 폴더의 products_with_supplier_info.xlsx 사용"

    def load_config(self):
        """마지막에 선택했던 공급사 파일 경로를 불러옵니다."""
        try:
            if self.config_path.exists():
                with open(self.config_path, "r", encoding="utf-8") as f:
                    config = json.load(f)
                supplier_path = clean_text(config.get("supplier_path", ""))
                if supplier_path and Path(supplier_path).exists():
                    self.supplier_path = Path(supplier_path)
        except Exception:
            self.supplier_path = None

    def save_config(self):
        """현재 선택된 공급사 파일 경로를 저장합니다."""
        try:
            config = {
                "supplier_path": str(self.supplier_path) if self.supplier_path else ""
            }
            self.config_path.parent.mkdir(parents=True, exist_ok=True)
            with open(self.config_path, "w", encoding="utf-8") as f:
                json.dump(config, f, ensure_ascii=False, indent=2)
        except Exception:
            pass

    def choose_supplier_file_and_reload(self):
        selected = filedialog.askopenfilename(
            title="supplier_detail_result.xlsx 파일 선택",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
        )
        if selected:
            self.supplier_path = Path(selected)
            self.save_config()
            self.path_label.configure(text=self._supplier_path_text())
            self.load_data()

    def load_data(self):
        try:
            self.status_var.set("데이터 로딩 중...")
            self.update_idletasks()

            # 상품 파일은 프로그램 폴더 기준 ../data/ 에서 고정으로 찾습니다.
            if not self.join_path.exists():
                messagebox.showerror(
                    "필수 파일 없음",
                    f"{DATA_DIR}/ 폴더에 상품 파일이 없습니다.\n\n"
                    f"현재 확인 위치:\n{self.join_path}\n\n"
                    f"products_with_supplier_info.xlsx 파일을 프로그램 폴더 기준 {DATA_DIR}/ 안에 넣어주세요."
                )
                self.status_var.set("상품 파일 없음")
                return

            # 공급사 파일은 사용자가 선택합니다. 마지막 선택 경로가 있으면 자동 사용합니다.
            if self.supplier_path is None or not self.supplier_path.exists():
                selected = filedialog.askopenfilename(
                    title="supplier_detail_result.xlsx 파일 선택",
                    filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
                )
                if not selected:
                    self.status_var.set("공급사 파일을 선택하지 않았습니다.")
                    return
                self.supplier_path = Path(selected)
                self.save_config()

            self.path_label.configure(text=self._supplier_path_text())

            self.supplier_df = read_excel_fast(self.supplier_path)
            self.goods_df = read_excel_fast(self.join_path)
            self.contact_status_df = self._read_or_create_contact_status()
            self.assignment_df = self.contact_status_df.copy()  # 기존 코드 호환용

            self._ensure_base_columns()
            self._prepare_supplier_list()
            self._refresh_filter_values()
            self.refresh_supplier_list()
            self.clear_all_detail_views()
            self.force_layout()
            self.status_var.set(
                f"로딩 완료: 공급사 {len(self.supplier_list_df):,}개 / "
                f"상품 {len(self.goods_df):,}건 / {self._contact_summary_text()}"
            )
        except Exception as e:
            self.status_var.set("로딩 실패")
            messagebox.showerror("오류", f"데이터 로딩 중 오류가 발생했습니다.\n\n{e}\n\n{traceback.format_exc()}")

    def _read_optional(self, path: Path) -> "pd.DataFrame":
        if path.exists():
            return read_excel_fast(path)
        return pd.DataFrame()

    def _read_or_create_contact_status(self) -> "pd.DataFrame":
        """공급사_연락상태.xlsx를 읽습니다. 없으면 빈 구조로 시작합니다."""
        if self.contact_status_path.exists():
            df = read_excel_fast(self.contact_status_path)
        else:
            df = pd.DataFrame(columns=CONTACT_STATUS_COLUMNS)

        for col in CONTACT_STATUS_COLUMNS:
            if col not in df.columns:
                df[col] = ""

        df = normalize_df(df)
        df["업체코드_KEY"] = df["업체코드"].map(norm_key)
        df["업체명_KEY"] = df["업체명"].map(norm_key)
        df["공급사_KEY"] = df["업체코드_KEY"] + "|" + df["업체명_KEY"]
        return df

    def save_contact_status_file(self):
        """연락상태 파일을 프로그램 폴더에 저장합니다."""
        try:
            out = self.contact_status_df.copy()
            for col in CONTACT_STATUS_COLUMNS:
                if col not in out.columns:
                    out[col] = ""
            out = out[CONTACT_STATUS_COLUMNS].copy()
            out.to_excel(self.contact_status_path, index=False)
        except Exception as e:
            messagebox.showerror("오류", f"연락상태 저장 실패: {e}")

    def _ensure_base_columns(self):
        for df_name in ["supplier_df", "goods_df", "contact_status_df"]:
            df = getattr(self, df_name)
            if df is None or df.empty:
                continue
            for col in ["업체코드", "업체명"]:
                if col not in df.columns:
                    df[col] = ""
            df["업체코드_KEY"] = df["업체코드"].map(norm_key)
            df["업체명_KEY"] = df["업체명"].map(norm_key)
            df["공급사_KEY"] = df["업체코드_KEY"] + "|" + df["업체명_KEY"]
            setattr(self, df_name, df)

        # 상품 필수 컬럼 보정
        if "카테고리" not in self.goods_df.columns:
            self.goods_df["카테고리"] = ""
        if "상품명" not in self.goods_df.columns:
            self.goods_df["상품명"] = ""
        for col in PRODUCT_COLUMNS:
            if col not in self.goods_df.columns:
                self.goods_df[col] = ""

        self.goods_df["카테고리_표시"] = self.goods_df["카테고리"].map(category_display)
        self.goods_df["대분류_표시"] = self.goods_df["카테고리"].map(major_display)
        self.goods_df["카테고리_KEY"] = self.goods_df["카테고리_표시"].map(norm_key)
        self.goods_df["대분류_KEY"] = self.goods_df["대분류_표시"].map(norm_key)
        self.goods_df["진열_표시"] = self.goods_df["진열"].map(display_status)
        self.goods_df["진열_KEY"] = self.goods_df["진열_표시"].map(norm_key)

    def _prepare_supplier_list(self):
        # 공급사 목록은 사용자가 선택한 공급사 파일 기준. 단, 상품 파일에만 있는 공급사는 맨 뒤에 추가.
        supplier = self.supplier_df.copy()
        supplier["자료출처"] = "공급사정보"

        goods_suppliers = self.goods_df[["업체코드", "업체명", "공급사_KEY", "업체코드_KEY", "업체명_KEY"]].drop_duplicates().copy()
        missing = goods_suppliers[~goods_suppliers["공급사_KEY"].isin(set(supplier["공급사_KEY"]))].copy()
        if not missing.empty:
            missing["자료출처"] = "상품파일만 존재"
            for col in supplier.columns:
                if col not in missing.columns:
                    missing[col] = ""
            supplier = pd.concat([supplier, missing[supplier.columns]], ignore_index=True)

        # 상품 파일 기준 집계값
        goods_pair_counts = self.goods_df.groupby("공급사_KEY", dropna=False).agg(
            조인상품수=("상품명", "size"),
            카테고리수=("카테고리_표시", lambda s: int(s.map(clean_text).replace("", pd.NA).dropna().nunique())),
        ).reset_index()
        goods_code_counts = self.goods_df.groupby("업체코드_KEY", dropna=False).agg(
            코드기준상품수=("상품명", "size"),
            코드기준카테고리수=("카테고리_표시", lambda s: int(s.map(clean_text).replace("", pd.NA).dropna().nunique())),
        ).reset_index()

        supplier = supplier.merge(goods_pair_counts, on="공급사_KEY", how="left")
        supplier = supplier.merge(goods_code_counts, on="업체코드_KEY", how="left")
        supplier["조인상품수"] = supplier["조인상품수"].fillna(supplier["코드기준상품수"]).fillna(0).astype(int)
        supplier["카테고리수"] = supplier["카테고리수"].fillna(supplier["코드기준카테고리수"]).fillna(0).astype(int)

        # 화면의 상품수는 상품 파일 기준으로 통일
        supplier["상품수"] = supplier["조인상품수"]

        # 연락상태 기본 컬럼
        for col in ["담당자", "연락여부", "연락일자", "연락수단", "연락결과", "재연락예정일", "메모", "최종수정일시"]:
            if col not in supplier.columns:
                supplier[col] = ""

        # 연락상태 파일이 있으면 공급사 목록에 붙임
        if self.contact_status_df is not None and not self.contact_status_df.empty:
            status_df = self.contact_status_df.copy()
            for col in CONTACT_STATUS_COLUMNS:
                if col not in status_df.columns:
                    status_df[col] = ""
            if "공급사_KEY" not in status_df.columns:
                status_df["업체코드_KEY"] = status_df["업체코드"].map(norm_key)
                status_df["업체명_KEY"] = status_df["업체명"].map(norm_key)
                status_df["공급사_KEY"] = status_df["업체코드_KEY"] + "|" + status_df["업체명_KEY"]

            status_df = status_df.drop_duplicates("공급사_KEY", keep="last")
            merge_cols = [
                "공급사_KEY", "담당자", "연락여부", "연락일자", "연락수단",
                "연락결과", "재연락예정일", "메모", "최종수정일시"
            ]
            supplier = supplier.merge(status_df[merge_cols], on="공급사_KEY", how="left", suffixes=("", "_상태"))
            for col in ["담당자", "연락여부", "연락일자", "연락수단", "연락결과", "재연락예정일", "메모", "최종수정일시"]:
                new_col = f"{col}_상태"
                if new_col in supplier.columns:
                    supplier[col] = supplier[new_col].where(supplier[new_col].map(clean_text).ne(""), supplier[col])
                    supplier = supplier.drop(columns=[new_col])

        supplier["연락여부"] = supplier["연락여부"].map(clean_text).replace("", "미연락")
        supplier["연락"] = supplier["연락여부"].map(self._contact_mark)

        supplier["상품수_num"] = supplier["상품수"].map(safe_int)
        supplier["카테고리수_num"] = supplier["카테고리수"].map(safe_int)
        supplier = supplier.sort_values(["상품수_num", "업체명"], ascending=[False, True])
        self.supplier_list_df = supplier.reset_index(drop=True)

    def _refresh_filter_values(self):
        # 담당자/상태 필터 UI는 사용하지 않지만, 내부 변수는 유지
        if not hasattr(self, "owner_filter_var"):
            self.owner_filter_var = tk.StringVar(value="전체")
        if not hasattr(self, "status_filter_var"):
            self.status_filter_var = tk.StringVar(value="전체")
        if hasattr(self, "contact_filter_var"):
            self.contact_filter_var.set("전체")
        self.owner_filter_var.set("전체")
        self.status_filter_var.set("전체")

    # ---------------- Supplier list/detail ----------------
    def _contact_mark(self, status: str) -> str:
        status = clean_text(status)
        if status == "연락완료":
            return "☑"
        if status == "재연락필요":
            return "↻"
        if status == "연락불가":
            return "×"
        return "☐"

    def _contact_tag(self, status: str) -> str:
        status = clean_text(status)
        if status == "연락완료":
            return "contact_done"
        if status == "재연락필요":
            return "contact_retry"
        if status == "연락불가":
            return "contact_fail"
        return "contact_none"

    def _count_contact_status(self, status: str) -> int:
        if self.supplier_list_df.empty or "연락여부" not in self.supplier_list_df.columns:
            return 0
        return int(self.supplier_list_df["연락여부"].map(clean_text).eq(status).sum())

    def _contact_summary_text(self) -> str:
        if self.supplier_list_df.empty or "연락여부" not in self.supplier_list_df.columns:
            return "연락상태 없음"
        total = len(self.supplier_list_df)
        done = self._count_contact_status("연락완료")
        none = self._count_contact_status("미연락")
        retry = self._count_contact_status("재연락필요")
        fail = self._count_contact_status("연락불가")
        return f"전체 {total:,}개 / 연락완료 {done:,}개 / 미연락 {none:,}개 / 재연락필요 {retry:,}개 / 연락불가 {fail:,}개"

    def refresh_supplier_list(self):
        self._clear_tree(self.supplier_tree)
        if self.supplier_list_df.empty:
            return

        keyword = self.supplier_search_var.get().strip().lower()
        owner = self.owner_filter_var.get() if hasattr(self, "owner_filter_var") else "전체"
        status = self.status_filter_var.get() if hasattr(self, "status_filter_var") else "전체"
        contact_filter = self.contact_filter_var.get() if hasattr(self, "contact_filter_var") else "전체"

        df = self.supplier_list_df.copy()
        if keyword:
            target = (
                df.get("업체명", "").astype(str) + " " +
                df.get("업체코드", "").astype(str) + " " +
                df.get("전화번호", "").astype(str) + " " +
                df.get("휴대폰번호", "").astype(str) + " " +
                df.get("이메일", "").astype(str)
            ).str.lower()
            df = df[target.str.contains(keyword, na=False)]
        if owner and owner != "전체":
            df = df[df.get("담당자", "").astype(str).eq(owner)]
        if status and status != "전체" and "검토상태" in df.columns:
            df = df[df.get("검토상태", "").astype(str).eq(status)]
        if contact_filter and contact_filter != "전체":
            df = df[df.get("연락여부", "").astype(str).eq(contact_filter)]

        for idx, row in df.iterrows():
            values = []
            for col in SUPPLIER_LIST_COLUMNS:
                if col == "상품수":
                    values.append(f"{safe_int(row.get('상품수', 0)):,}")
                else:
                    values.append(clean_text(row.get(col, "")))
            iid = f"supplier_{idx}"
            contact_status = clean_text(row.get("연락여부", "미연락")) or "미연락"
            self.supplier_tree.insert("", tk.END, iid=iid, values=values, tags=(self._contact_tag(contact_status),))
        self.status_var.set(f"공급사 목록 {len(df):,}개 표시 / {self._contact_summary_text()}")

    def on_supplier_select(self, event=None):
        item_id = self._selected_item(self.supplier_tree)
        if not item_id:
            return
        try:
            idx = int(str(item_id).replace("supplier_", "", 1))
            row = self.supplier_list_df.loc[idx]
        except Exception:
            messagebox.showerror("오류", "선택한 공급사 정보를 찾을 수 없습니다.")
            return

        self.selected_supplier_code = clean_text(row.get("업체코드", ""))
        self.selected_supplier_name = clean_text(row.get("업체명", ""))
        self.current_category_filter = ""
        self.current_major_filter = ""
        self.product_display_count = self.product_display_limit
        self.product_search_var.set("")
        self.current_display_status_filter = "전체"
        if hasattr(self, "display_status_filter_var"):
            self.display_status_filter_var.set("전체")
        self.category_filter_label.configure(text="필터 없음")

        self.show_supplier_detail()
        self.show_major_summary()
        self.show_category_summary()
        self.prepare_product_df()
        self.refresh_product_list()

    def _selected_supplier_row(self) -> "pd.Series | None":
        if self.supplier_list_df.empty:
            return None
        code_key = norm_key(self.selected_supplier_code)
        name_key = norm_key(self.selected_supplier_name)
        hit = self.supplier_list_df[
            (self.supplier_list_df["업체코드_KEY"].eq(code_key)) &
            (self.supplier_list_df["업체명_KEY"].eq(name_key))
        ]
        if hit.empty and code_key:
            hit = self.supplier_list_df[self.supplier_list_df["업체코드_KEY"].eq(code_key)]
        if hit.empty and name_key:
            hit = self.supplier_list_df[self.supplier_list_df["업체명_KEY"].eq(name_key)]
        if hit.empty:
            return None
        return hit.iloc[0]
        
    def show_supplier_detail(self):
        row = self._selected_supplier_row()

        self.info_text.configure(state=tk.NORMAL)
        self.info_text.delete("1.0", tk.END)

        # 오른쪽 단 시작 위치 고정
        # 숫자를 키우면 [상품/연락 정보]가 더 오른쪽으로 이동합니다.
        RIGHT_TAB = 400

        self.info_text.configure(
            font=("맑은 고딕", 10),
            foreground="#111111",
            background="#FFFFFF",
            spacing1=2,
            spacing2=1,
            spacing3=6,
            padx=14,
            pady=12,
            wrap=tk.WORD,
            tabs=(RIGHT_TAB,),
        )
        
        self.info_text.tag_configure(
            "company",
            foreground="#000000",
            font=("맑은 고딕", 12, "bold"),
        )
        self.info_text.tag_configure(
            "summary",
            foreground="#000000",
            font=("맑은 고딕", 10, "bold"),
        )
        self.info_text.tag_configure(
            "section",
            foreground="#1F4E79",
            font=("맑은 고딕", 10, "bold"),
        )
        self.info_text.tag_configure(
            "label",
            foreground="#333333",
            font=("맑은 고딕", 10, "bold"),
        )
        self.info_text.tag_configure(
            "value",
            foreground="#111111",
            font=("맑은 고딕", 10),
        )
        self.info_text.tag_configure(
            "empty",
            foreground="#777777",
            font=("맑은 고딕", 10),
        )
        self.info_text.tag_configure(
            "compact_line",
            spacing1=0,
            spacing2=0,
            spacing3=0,
        )


        if row is None:
            self.info_text.insert(tk.END, "표시할 공급사 정보가 없습니다.", "empty")
            self.info_text.configure(state=tk.DISABLED)
            return

        def get_field_lines(fields):
            lines = []
            for col in fields:
                if col in row.index:
                    val = clean_text(row.get(col, ""))
                    if val:
                        label = FIELD_LABELS.get(col, col)
                        lines.append((label, val))
            return lines

        left_lines = get_field_lines(SUPPLIER_LEFT_FIELDS)
        right_lines = get_field_lines(SUPPLIER_RIGHT_FIELDS)

        if not left_lines and not right_lines:
            self.info_text.insert(tk.END, "표시할 공급사 정보가 없습니다.", "empty")
            self.info_text.configure(state=tk.DISABLED)
            return

        # 상단 요약
        company_name = clean_text(row.get("업체명", ""))
        product_count = safe_int(row.get("상품수", 0))
        category_count = safe_int(row.get("카테고리수", 0))
        contact_status = clean_text(row.get("연락여부", "미연락")) or "미연락"
        contact_date = clean_text(row.get("연락일자", ""))

        self.info_text.insert(tk.END, company_name, "company")
        self.info_text.insert(tk.END, "\t")
        self.info_text.insert(tk.END, f"상품수 : {product_count:,}", "summary")
        self.info_text.insert(tk.END, "      ")
        self.info_text.insert(tk.END, f"카테고리수 : {category_count:,}", "summary")
        self.info_text.insert(tk.END, "\n\n")

        # 섹션 제목
        self.info_text.insert(tk.END, "[사업자 정보]", "section")
        self.info_text.insert(tk.END, "\t")
        self.info_text.insert(tk.END, "[연락 정보]\n", "section")

        # 긴 값은 쉼표 기준으로 줄바꿈
        def wrap_long_value(text, limit=28):
            text = clean_text(text)

            if len(text) <= limit:
                return [text]

            parts = [p.strip() for p in text.split(",")]
            lines = []
            cur = ""

            for part in parts:
                if not part:
                    continue

                candidate = part if not cur else f"{cur}, {part}"

                if len(candidate) <= limit:
                    cur = candidate
                else:
                    if cur:
                        lines.append(cur)
                    cur = part

            if cur:
                lines.append(cur)

            return lines or [text]


        max_rows = max(len(left_lines), len(right_lines))

        for i in range(max_rows):
            # 왼쪽 단
            left_label = ""
            extra_lines = []

            if i < len(left_lines):
                left_label, left_value = left_lines[i]

                # 주소는 우편번호와 상세주소를 줄바꿈해서 표시
                if left_label == "주소":
                    address_text = clean_text(left_value)

                    # 예: "우) 06231 서울 강남구 ..." → "우) 06231" / "서울 강남구 ..."
                    m = re.match(r"^(우\)\s*\d{5})\s+(.*)$", address_text)

                    self.info_text.insert(tk.END, f"{left_label} : ", ("label", "compact_line"))

                    if m:
                        postcode = m.group(1)
                        detail_addr = m.group(2)

                        self.info_text.insert(tk.END, f"{postcode}\n", ("value", "compact_line"))

                        # '주소 : ' 길이만큼 들여쓰기해서 서울이 우) 아래에서 시작
                        indent = " " * (len(f"{left_label} : ") + 3)
                        self.info_text.insert(tk.END, indent, "compact_line")
                        self.info_text.insert(tk.END, f"{detail_addr}", ("value", "compact_line"))
                    else:
                        self.info_text.insert(tk.END, address_text, ("value", "compact_line"))

                    self.info_text.insert(tk.END, "\n", "compact_line")
                    continue

                else:
                    label_text = f"{left_label} : "
                    wrapped_lines = wrap_long_value(left_value, limit=32)

                    self.info_text.insert(tk.END, label_text, "label")
                    self.info_text.insert(tk.END, wrapped_lines[0], ("value", "compact_line"))

                    # 길어서 다음 줄로 내려갈 내용은 오른쪽 단 출력 후 아래에 표시
                    extra_lines = wrapped_lines[1:]

            # 주소 줄은 길어서 오른쪽 단을 붙이지 않음
            is_address_line = left_label == "주소"

            # 오른쪽 단
            if not is_address_line and i < len(right_lines):
                right_label, right_value = right_lines[i]
                self.info_text.insert(tk.END, "\t")
                self.info_text.insert(tk.END, f"{right_label} : ", "label")
                self.info_text.insert(tk.END, f"{right_value}", "value")

            if extra_lines:
                self.info_text.insert(tk.END, "\n", "compact_line")
            else:
                self.info_text.insert(tk.END, "\n")

            # 왼쪽 값이 길면, 오른쪽 단 아래가 아니라 왼쪽 값 아래로만 이어서 표시
            if extra_lines:
                indent = " " * (len(f"{left_label} : ") + 3)

                for idx, extra_line in enumerate(extra_lines):
                    self.info_text.insert(tk.END, indent, "compact_line")
                    self.info_text.insert(tk.END, extra_line, ("value", "compact_line"))

                    # 추가 줄들끼리는 좁게, 마지막 줄 다음은 일반 간격
                    if idx == len(extra_lines) - 1:
                        self.info_text.insert(tk.END, "\n")
                    else:
                        self.info_text.insert(tk.END, "\n", "compact_line")

    # ---------------- Major/category/product views ----------------
    def _supplier_goods_rows(self, code: str, name: str) -> "pd.DataFrame":
        if self.goods_df.empty:
            return pd.DataFrame()
        code_key = norm_key(code)
        name_key = norm_key(name)
        df = self.goods_df

        if code_key and name_key:
            hit = df[(df["업체코드_KEY"].eq(code_key)) & (df["업체명_KEY"].eq(name_key))]
            if not hit.empty:
                return hit.copy()
        if code_key:
            hit = df[df["업체코드_KEY"].eq(code_key)]
            if not hit.empty:
                return hit.copy()
        if name_key:
            hit = df[df["업체명_KEY"].eq(name_key)]
            if not hit.empty:
                return hit.copy()
        return pd.DataFrame()

    def _exclude_uncategorized_if_needed(self, gdf: "pd.DataFrame") -> "pd.DataFrame":
        if SHOW_UNCATEGORIZED:
            return gdf
        return gdf[gdf["카테고리"].map(clean_text).ne("")]

    def show_major_summary(self):
        self._clear_tree(self.major_tree)
        gdf = self._supplier_goods_rows(self.selected_supplier_code, self.selected_supplier_name)
        if gdf.empty:
            return
        gdf = self._exclude_uncategorized_if_needed(gdf)
        if gdf.empty:
            return
        summary = gdf["대분류_표시"].map(clean_text).replace("", "미분류").value_counts().reset_index()
        summary.columns = ["대분류", "상품수"]
        for _, r in summary.iterrows():
            self.major_tree.insert("", tk.END, values=(clean_text(r["대분류"]), f"{safe_int(r['상품수']):,}"))

    def show_category_summary(self):
        self._clear_tree(self.category_tree)
        gdf = self._supplier_goods_rows(self.selected_supplier_code, self.selected_supplier_name)
        if gdf.empty:
            return
        gdf = self._exclude_uncategorized_if_needed(gdf)
        if gdf.empty:
            return
        vc = gdf["카테고리_표시"].map(clean_text).replace("", "미분류").value_counts().reset_index()
        vc.columns = ["카테고리", "개수"]
        # for _, r in vc.iterrows():
        #     category = clean_text(r["카테고리"])
        #     parts = [p.strip() for p in category.split(">")]
        #     parts += [""] * (4 - len(parts))
        #     self.category_tree.insert("", tk.END, values=(category, f"{safe_int(r['개수']):,}", parts[0], parts[1], parts[2], parts[3]))
        for _, r in vc.iterrows():
            category = clean_text(r["카테고리"])
            self.category_tree.insert(
                "",
                tk.END,
                values=(category, f"{safe_int(r['개수']):,}")
            )

    def prepare_product_df(self):
        df = self._supplier_goods_rows(self.selected_supplier_code, self.selected_supplier_name)
        if df.empty:
            self.current_product_df = pd.DataFrame(columns=PRODUCT_COLUMNS)
            return
        df = df.copy()
        # 화면에서는 카테고리/진열상태 빈 값을 미분류로 보여줌
        df["카테고리"] = df["카테고리_표시"]
        df["진열"] = df["진열_표시"]
        for col in PRODUCT_COLUMNS:
            if col not in df.columns:
                df[col] = ""
        self.current_product_df = df[PRODUCT_COLUMNS + ["카테고리_KEY", "대분류_KEY", "진열_KEY"]].copy()

    def refresh_product_list(self):
        self._clear_tree(self.product_tree)
        if self.current_product_df.empty:
            self.current_product_filtered_df = pd.DataFrame(columns=PRODUCT_COLUMNS)
            self.product_count_var.set("상품 0건 — 상품 파일에서 매칭되는 행이 없습니다")
            return

        df = self.current_product_df.copy()
        if not SHOW_UNCATEGORIZED:
            df = df[df["카테고리"].map(clean_text).ne("미분류")]

        if self.current_category_filter:
            cat_key = norm_key(self.current_category_filter)
            df = df[df["카테고리_KEY"].eq(cat_key)]
        if self.current_major_filter:
            major_key = norm_key(self.current_major_filter)
            df = df[df["대분류_KEY"].eq(major_key)]

        display_status_filter = clean_text(
            self.display_status_filter_var.get() if hasattr(self, "display_status_filter_var") else self.current_display_status_filter
        )
        if display_status_filter and display_status_filter != "전체":
            df = df[df["진열_KEY"].eq(norm_key(display_status_filter))]

        keyword = self.product_search_var.get().strip().lower()
        if keyword:
            target = (
                df.get("상품명", "").astype(str) + " " +
                df.get("상품코드", "").astype(str) + " " +
                df.get("카테고리", "").astype(str)
            ).str.lower()
            df = df[target.str.contains(keyword, na=False)]

        self.current_product_filtered_df = df[PRODUCT_COLUMNS].copy()
        total = len(self.current_product_filtered_df)
        show_count = min(total, max(self.product_display_count, self.product_display_limit))
        for _, row in self.current_product_filtered_df.head(show_count).iterrows():
            values = [clean_text(row.get(col, "")) for col in PRODUCT_COLUMNS]
            self.product_tree.insert("", tk.END, values=values)
        suffix = "" if total <= show_count else f" / 화면 표시 {show_count:,}건, 더 보기 가능"
        self.product_count_var.set(f"상품 {total:,}건{suffix}")
        self.status_var.set(f"{self.selected_supplier_name} 상품 {total:,}건 조회")

    def on_display_status_filter_change(self, event=None):
        self.current_display_status_filter = clean_text(self.display_status_filter_var.get()) or "전체"
        self.product_display_count = self.product_display_limit
        self.refresh_product_list()

    def on_category_select(self, event=None):
        item_id = self._selected_item(self.category_tree)
        if not item_id:
            return
        values = self.category_tree.item(item_id, "values")
        category = clean_text(values[0])
        if not category:
            return
        self.current_category_filter = category
        self.current_major_filter = ""
        self.product_display_count = self.product_display_limit
        self.category_filter_label.configure(text=f"카테고리 필터: {category}")
        self.refresh_product_list()

    def on_major_select(self, event=None):
        item_id = self._selected_item(self.major_tree)
        if not item_id:
            return
        values = self.major_tree.item(item_id, "values")
        major = clean_text(values[0])
        if not major:
            return
        self.current_major_filter = major
        self.current_category_filter = ""
        self.product_display_count = self.product_display_limit
        self.category_filter_label.configure(text=f"대분류 필터: {major}")
        self.refresh_product_list()

    def clear_category_filter(self):
        self.current_category_filter = ""
        self.current_major_filter = ""
        self.product_display_count = self.product_display_limit
        self.category_filter_label.configure(text="필터 없음")
        self.refresh_product_list()

    def show_more_products(self):
        self.product_display_count += self.product_display_limit
        self.refresh_product_list()

    # ---------------- Contact status ----------------
    def update_contact_status(self, status: str):
        """선택한 공급사의 연락상태를 저장하고 화면에 표시합니다."""
        row = self._selected_supplier_row()
        if row is None:
            messagebox.showinfo("안내", "공급사를 먼저 선택하세요.")
            return

        supplier_code = clean_text(row.get("업체코드", ""))
        supplier_name = clean_text(row.get("업체명", ""))
        code_key = norm_key(supplier_code)
        name_key = norm_key(supplier_name)
        supplier_key = code_key + "|" + name_key

        now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        today = datetime.now().strftime("%Y-%m-%d")

        if self.contact_status_df is None or self.contact_status_df.empty:
            self.contact_status_df = pd.DataFrame(columns=CONTACT_STATUS_COLUMNS)

        for col in CONTACT_STATUS_COLUMNS:
            if col not in self.contact_status_df.columns:
                self.contact_status_df[col] = ""

        if "공급사_KEY" not in self.contact_status_df.columns:
            self.contact_status_df["업체코드_KEY"] = self.contact_status_df["업체코드"].map(norm_key)
            self.contact_status_df["업체명_KEY"] = self.contact_status_df["업체명"].map(norm_key)
            self.contact_status_df["공급사_KEY"] = self.contact_status_df["업체코드_KEY"] + "|" + self.contact_status_df["업체명_KEY"]

        hit = self.contact_status_df["공급사_KEY"].eq(supplier_key)

        if hit.any():
            idx = self.contact_status_df[hit].index[-1]
        else:
            new_row = {
                "업체코드": supplier_code,
                "업체명": supplier_name,
                "담당자": clean_text(row.get("담당자", "")),
                "연락여부": "",
                "연락일자": "",
                "연락수단": "",
                "연락결과": "",
                "재연락예정일": "",
                "메모": clean_text(row.get("메모", "")),
                "최종수정일시": "",
                "업체코드_KEY": code_key,
                "업체명_KEY": name_key,
                "공급사_KEY": supplier_key,
            }
            self.contact_status_df = pd.concat([self.contact_status_df, pd.DataFrame([new_row])], ignore_index=True)
            idx = self.contact_status_df.index[-1]

        self.contact_status_df.loc[idx, "업체코드"] = supplier_code
        self.contact_status_df.loc[idx, "업체명"] = supplier_name
        self.contact_status_df.loc[idx, "연락여부"] = status
        self.contact_status_df.loc[idx, "최종수정일시"] = now

        if status == "연락완료":
            self.contact_status_df.loc[idx, "연락일자"] = today
            self.contact_status_df.loc[idx, "연락결과"] = "연락완료"
        elif status == "재연락필요":
            if not clean_text(self.contact_status_df.loc[idx, "연락일자"]):
                self.contact_status_df.loc[idx, "연락일자"] = today
            self.contact_status_df.loc[idx, "연락결과"] = "재연락필요"
        elif status == "연락불가":
            if not clean_text(self.contact_status_df.loc[idx, "연락일자"]):
                self.contact_status_df.loc[idx, "연락일자"] = today
            self.contact_status_df.loc[idx, "연락결과"] = "연락불가"
        elif status == "미연락":
            self.contact_status_df.loc[idx, "연락일자"] = ""
            self.contact_status_df.loc[idx, "연락결과"] = ""

        self.save_contact_status_file()

        # 선택 공급사 유지용
        self.selected_supplier_code = supplier_code
        self.selected_supplier_name = supplier_name

        self._prepare_supplier_list()
        self.refresh_supplier_list()
        self.show_supplier_detail()

        self.status_var.set(f"{supplier_name} → {status} 처리 완료 / {self._contact_summary_text()}")

    # ---------------- Actions ----------------
    def open_selected_product_link(self):
        item_id = self._selected_item(self.product_tree)
        if not item_id:
            messagebox.showinfo("안내", "상품을 선택하세요.")
            return
        values = self.product_tree.item(item_id, "values")
        link_candidates = []
        if "상품링크" in PRODUCT_COLUMNS:
            link_candidates.append(values[PRODUCT_COLUMNS.index("상품링크")])
        if "이미지URL" in PRODUCT_COLUMNS:
            link_candidates.append(values[PRODUCT_COLUMNS.index("이미지URL")])
        url = next((clean_text(u) for u in link_candidates if clean_text(u).startswith("http")), "")
        if not url:
            messagebox.showinfo("안내", "선택한 상품에 열 수 있는 링크가 없습니다.")
            return
        webbrowser.open(url)

    def export_current_products(self):
        if self.current_product_filtered_df.empty:
            messagebox.showinfo("안내", "저장할 상품 목록이 없습니다.")
            return
        default_name = self._safe_filename(f"{self.selected_supplier_name}_상품목록_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx")
        path = filedialog.asksaveasfilename(title="상품목록 저장", defaultextension=".xlsx", initialfile=default_name, filetypes=[("Excel files", "*.xlsx")])
        if not path:
            return
        try:
            self.current_product_filtered_df.to_excel(path, index=False)
            messagebox.showinfo("완료", f"저장 완료\n{path}")
        except Exception as e:
            messagebox.showerror("오류", f"저장 실패: {e}")

    def export_supplier_summary(self):
        if not self.selected_supplier_name:
            messagebox.showinfo("안내", "공급사를 선택하세요.")
            return
        default_name = self._safe_filename(f"{self.selected_supplier_name}_공급사요약_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx")
        path = filedialog.asksaveasfilename(title="공급사 요약 저장", defaultextension=".xlsx", initialfile=default_name, filetypes=[("Excel files", "*.xlsx")])
        if not path:
            return
        try:
            info_df = self._supplier_info_as_df()
            major_df = self._tree_to_df(self.major_tree)
            category_df = self._tree_to_df(self.category_tree)
            products_df = self.current_product_filtered_df if not self.current_product_filtered_df.empty else self.current_product_df[PRODUCT_COLUMNS]
            with pd.ExcelWriter(path, engine="openpyxl") as writer:
                info_df.to_excel(writer, index=False, sheet_name="공급사정보")
                major_df.to_excel(writer, index=False, sheet_name="대분류요약")
                category_df.to_excel(writer, index=False, sheet_name="카테고리요약")
                products_df.to_excel(writer, index=False, sheet_name="상품목록")
            messagebox.showinfo("완료", f"저장 완료\n{path}")
        except Exception as e:
            messagebox.showerror("오류", f"저장 실패: {e}")

    def export_assignment_template(self):
        """현재 공급사 목록과 연락상태를 엑셀로 저장합니다."""
        if self.supplier_list_df.empty:
            messagebox.showinfo("안내", "저장할 공급사 목록이 없습니다.")
            return
        default_name = f"공급사_연락상태_목록_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
        path = filedialog.asksaveasfilename(
            title="연락상태 목록 저장",
            defaultextension=".xlsx",
            initialfile=default_name,
            filetypes=[("Excel files", "*.xlsx")],
        )
        if not path:
            return
        try:
            cols = [
                "연락", "연락여부", "연락일자", "연락수단", "연락결과", "재연락예정일", "담당자", "메모",
                "업체코드", "업체명", "상품수", "조인상품수", "카테고리수",
                "전화번호", "휴대폰번호", "이메일", "주소", "인쇄가능여부", "발주링크", "상품링크", "자료출처", "최종수정일시"
            ]
            out = self.supplier_list_df.copy()
            for col in cols:
                if col not in out.columns:
                    out[col] = ""
            out[cols].to_excel(path, index=False)
            messagebox.showinfo(
                "완료",
                f"연락상태 목록을 저장했습니다.\n\n{path}\n\n"
                f"실제 연락상태 누적 파일은 프로그램 폴더의 '{CONTACT_STATUS_FILE}'에 자동 저장됩니다."
            )
        except Exception as e:
            messagebox.showerror("오류", f"저장 실패: {e}")

    def _supplier_info_as_df(self) -> "pd.DataFrame":
        text = self.info_text.get("1.0", tk.END).strip()
        rows = []
        for line in text.splitlines():
            if ":" in line:
                k, v = line.split(":", 1)
                rows.append({"항목": k.strip(), "내용": v.strip()})
        return pd.DataFrame(rows)

    # ---------------- Utilities ----------------
    def _clear_tree(self, tree: ttk.Treeview):
        tree.delete(*tree.get_children())

    def _selected_item(self, tree: ttk.Treeview) -> str:
        selected = tree.selection()
        return selected[0] if selected else ""

    def _tree_to_df(self, tree: ttk.Treeview) -> "pd.DataFrame":
        cols = list(tree["columns"])
        rows = []
        for item in tree.get_children():
            values = list(tree.item(item, "values"))
            rows.append(dict(zip(cols, values)))
        return pd.DataFrame(rows)

    def _safe_filename(self, name: str) -> str:
        for ch in '\\/:*?"<>|':
            name = name.replace(ch, "_")
        return name

    def sort_treeview(self, tree: ttk.Treeview, col: str, reverse: bool):
        try:
            data = [(tree.set(k, col), k) for k in tree.get_children("")]
            def sort_key(item):
                value = clean_text(item[0]).replace(",", "")
                try:
                    return float(value)
                except Exception:
                    return value
            data.sort(key=sort_key, reverse=reverse)
            for index, (_, k) in enumerate(data):
                tree.move(k, "", index)
            tree.heading(col, command=lambda: self.sort_treeview(tree, col, not reverse))
        except Exception:
            pass

    def clear_all_detail_views(self):
        for tree in [self.major_tree, self.category_tree, self.product_tree]:
            self._clear_tree(tree)
        self.current_product_df = pd.DataFrame(columns=PRODUCT_COLUMNS)
        self.current_product_filtered_df = pd.DataFrame(columns=PRODUCT_COLUMNS)
        self.product_count_var.set("상품 0건")
        self.current_display_status_filter = "전체"
        if hasattr(self, "display_status_filter_var"):
            self.display_status_filter_var.set("전체")
        self.category_filter_label.configure(text="필터 없음")
        self.info_text.configure(state=tk.NORMAL)
        self.info_text.delete("1.0", tk.END)
        self.info_text.insert(tk.END, "왼쪽에서 공급사를 선택하세요.")
        self.info_text.configure(state=tk.DISABLED)


def main():
    app = SupplierProductViewer()
    app.mainloop()


if __name__ == "__main__":
    main()
