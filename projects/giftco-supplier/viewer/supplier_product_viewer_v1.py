# -*- coding: utf-8 -*-
"""
공급사별 상품 카테고리/상품 조회 툴

필요 파일을 이 스크립트와 같은 폴더에 두고 실행하세요.
- products_with_supplier_info.xlsx       : 필수, 상품 + 공급사 정보 조인 파일
- 업체별_대분류.xlsx                  : 선택, 공급사별 대분류 집계 파일
- 업체별_카테고리.xlsx                : 선택, 공급사별 카테고리 집계 파일
- supplier_detail_result.xlsx           : 선택, 공급사 세부정보 파일

실행 전 설치:
    pip install pandas openpyxl

실행:
    python supplier_product_viewer_v1.py
"""

from __future__ import annotations

import os
import sys
import webbrowser
import traceback
from pathlib import Path
from datetime import datetime
import tkinter as tk
from tkinter import ttk, filedialog, messagebox
from tkinter.scrolledtext import ScrolledText

try:
    import pandas as pd
except ImportError:
    pd = None

try:
    from openpyxl import load_workbook
except ImportError:
    load_workbook = None


DATA_DIR = "../data"
REQUIRED_JOIN_FILE = f"{DATA_DIR}/products_with_supplier_info.xlsx"
OPTIONAL_MAJOR_FILE = f"{DATA_DIR}/업체별_대분류.xlsx"
OPTIONAL_CATEGORY_FILE = f"{DATA_DIR}/업체별_카테고리.xlsx"
OPTIONAL_SUPPLIER_FILE = f"{DATA_DIR}/supplier_detail_result.xlsx"

SUPPLIER_KEY_COLS = ["업체코드", "업체명"]
PRODUCT_COLUMNS = [
    "상품코드", "상품명", "카테고리", "기본수량", "공급가", "판매가1", "판매가7",
    "마진율", "진열", "노출", "최초등록일", "최근수정일", "상품링크"
]
SUPPLIER_INFO_COLUMNS = [
    "업체명", "업체코드", "상품수", "전화번호", "휴대폰번호", "팩스", "이메일", "주소",
    "상품등록일", "상품최근수정일", "인쇄가능여부", "한박스당수량", "한박스당배송비", "이미지사용",
    "발주링크", "상호명", "사업자번호", "대표자", "업태", "종목", "우편번호"
]


def clean_text(value) -> str:
    """NaN/None을 빈 문자열로 정리하고 앞뒤 공백을 제거합니다."""
    if value is None:
        return ""
    try:
        if pd is not None and pd.isna(value):
            return ""
    except Exception:
        pass
    return str(value).strip()


def normalize_df(df: "pd.DataFrame") -> "pd.DataFrame":
    """컬럼명/문자열 데이터를 조회하기 편한 형태로 정리합니다.

    대용량 엑셀을 빠르게 읽기 위해 셀 단위 map 대신 벡터화된 문자열 처리를 사용합니다.
    """
    df = df.copy()
    df.columns = ["" if c is None else str(c).strip() for c in df.columns]
    df = df.fillna("")
    for col in df.columns:
        if df[col].dtype == object:
            df[col] = df[col].astype(str).str.strip()
    return df


def read_excel_fast(path: Path) -> "pd.DataFrame":
    """openpyxl read_only 방식으로 첫 번째 시트를 빠르게 DataFrame으로 읽습니다."""
    if load_workbook is None:
        raise ImportError("openpyxl이 설치되어 있지 않습니다. pip install openpyxl 후 다시 실행하세요.")
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)
    try:
        headers = next(rows_iter)
    except StopIteration:
        return pd.DataFrame()
    headers = ["" if h is None else str(h).strip() for h in headers]
    data = list(rows_iter)
    return normalize_df(pd.DataFrame(data, columns=headers))


def first_existing_column(df: "pd.DataFrame", candidates: list[str]) -> str | None:
    for col in candidates:
        if col in df.columns:
            return col
    return None


class SupplierProductViewer(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("공급사별 상품 카테고리 조회 툴")
        self.geometry("1500x900")
        self.minsize(1200, 720)

        if pd is None:
            messagebox.showerror(
                "pandas 설치 필요",
                "이 프로그램은 pandas/openpyxl이 필요합니다.\n\n명령 프롬프트에서 아래 명령을 실행하세요.\n\npip install pandas openpyxl"
            )
            self.destroy()
            return

        self.base_dir = Path(getattr(sys, "_MEIPASS", Path(__file__).resolve().parent))
        self.data_dir = Path(__file__).resolve().parent

        self.goods_df = pd.DataFrame()
        self.supplier_df = pd.DataFrame()
        self.major_df = pd.DataFrame()
        self.category_df = pd.DataFrame()
        self.supplier_summary_df = pd.DataFrame()

        self.selected_supplier_name = ""
        self.selected_supplier_code = ""
        self.current_product_df = pd.DataFrame()
        self.current_product_filtered_df = pd.DataFrame()
        self.current_category_filter = ""

        self._setup_style()
        self._build_layout()
        self.after(100, self.load_data)

    # ---------------- UI ----------------
    def _setup_style(self):
        style = ttk.Style(self)
        try:
            style.theme_use("clam")
        except Exception:
            pass
        style.configure("Treeview", rowheight=24, font=("맑은 고딕", 9))
        style.configure("Treeview.Heading", font=("맑은 고딕", 9, "bold"))
        style.configure("TLabel", font=("맑은 고딕", 9))
        style.configure("TButton", font=("맑은 고딕", 9))
        style.configure("TLabelframe.Label", font=("맑은 고딕", 10, "bold"))

    def _build_layout(self):
        top_bar = ttk.Frame(self, padding=(8, 6))
        top_bar.pack(fill=tk.X)

        ttk.Label(top_bar, text="데이터 폴더:").pack(side=tk.LEFT)
        self.path_label = ttk.Label(top_bar, text=str(self.data_dir), foreground="#555555")
        self.path_label.pack(side=tk.LEFT, padx=(6, 12), fill=tk.X, expand=True)
        ttk.Button(top_bar, text="폴더 선택/재로드", command=self.choose_folder_and_reload).pack(side=tk.RIGHT, padx=(6, 0))
        ttk.Button(top_bar, text="전체 새로고침", command=self.load_data).pack(side=tk.RIGHT)

        self.status_var = tk.StringVar(value="준비 중")
        status = ttk.Label(self, textvariable=self.status_var, anchor="w", padding=(8, 4))
        status.pack(fill=tk.X)

        main = ttk.PanedWindow(self, orient=tk.HORIZONTAL)
        main.pack(fill=tk.BOTH, expand=True, padx=8, pady=8)

        # Left supplier panel
        left = ttk.Frame(main)
        main.add(left, weight=1)
        supplier_box = ttk.LabelFrame(left, text="공급사 목록", padding=6)
        supplier_box.pack(fill=tk.BOTH, expand=True)

        search_frame = ttk.Frame(supplier_box)
        search_frame.pack(fill=tk.X, pady=(0, 6))
        ttk.Label(search_frame, text="검색").pack(side=tk.LEFT)
        self.supplier_search_var = tk.StringVar()
        self.supplier_search_var.trace_add("write", lambda *args: self.refresh_supplier_list())
        supplier_entry = ttk.Entry(search_frame, textvariable=self.supplier_search_var)
        supplier_entry.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=5)
        ttk.Button(search_frame, text="초기화", command=lambda: self.supplier_search_var.set("")).pack(side=tk.RIGHT)

        self.supplier_tree = ttk.Treeview(
            supplier_box,
            columns=("업체코드", "업체명", "상품수", "카테고리수"),
            show="headings",
            selectmode="browse",
        )
        self._config_tree(self.supplier_tree, {
            "업체코드": 90,
            "업체명": 210,
            "상품수": 70,
            "카테고리수": 80,
        })
        self.supplier_tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        supplier_scroll = ttk.Scrollbar(supplier_box, orient=tk.VERTICAL, command=self.supplier_tree.yview)
        self.supplier_tree.configure(yscrollcommand=supplier_scroll.set)
        supplier_scroll.pack(side=tk.RIGHT, fill=tk.Y)
        self.supplier_tree.bind("<<TreeviewSelect>>", self.on_supplier_select)

        # Right panel
        right = ttk.PanedWindow(main, orient=tk.VERTICAL)
        main.add(right, weight=4)

        top_right = ttk.PanedWindow(right, orient=tk.HORIZONTAL)
        right.add(top_right, weight=2)

        info_box = ttk.LabelFrame(top_right, text="공급사 정보", padding=6)
        top_right.add(info_box, weight=1)
        self.info_text = ScrolledText(info_box, height=12, font=("맑은 고딕", 9), wrap=tk.WORD)
        self.info_text.pack(fill=tk.BOTH, expand=True)
        self.info_text.configure(state=tk.DISABLED)

        major_box = ttk.LabelFrame(top_right, text="대분류 요약", padding=6)
        top_right.add(major_box, weight=1)
        self.major_tree = ttk.Treeview(major_box, columns=("대분류", "상품수"), show="headings", selectmode="browse")
        self._config_tree(self.major_tree, {"대분류": 240, "상품수": 90})
        self.major_tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        major_scroll = ttk.Scrollbar(major_box, orient=tk.VERTICAL, command=self.major_tree.yview)
        self.major_tree.configure(yscrollcommand=major_scroll.set)
        major_scroll.pack(side=tk.RIGHT, fill=tk.Y)
        self.major_tree.bind("<<TreeviewSelect>>", self.on_major_select)

        cat_product_pane = ttk.PanedWindow(right, orient=tk.HORIZONTAL)
        right.add(cat_product_pane, weight=5)

        category_box = ttk.LabelFrame(cat_product_pane, text="카테고리 목록", padding=6)
        cat_product_pane.add(category_box, weight=2)
        cat_btn = ttk.Frame(category_box)
        cat_btn.pack(fill=tk.X, pady=(0, 6))
        ttk.Button(cat_btn, text="카테고리 필터 해제", command=self.clear_category_filter).pack(side=tk.LEFT)
        self.category_filter_label = ttk.Label(cat_btn, text="필터 없음", foreground="#555555")
        self.category_filter_label.pack(side=tk.LEFT, padx=8)

        self.category_tree = ttk.Treeview(
            category_box,
            columns=("카테고리", "개수", "대", "중", "소", "세"),
            show="headings",
            selectmode="browse",
        )
        self._config_tree(self.category_tree, {
            "카테고리": 360,
            "개수": 60,
            "대": 130,
            "중": 150,
            "소": 150,
            "세": 120,
        })
        self.category_tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        cat_scroll = ttk.Scrollbar(category_box, orient=tk.VERTICAL, command=self.category_tree.yview)
        self.category_tree.configure(yscrollcommand=cat_scroll.set)
        cat_scroll.pack(side=tk.RIGHT, fill=tk.Y)
        self.category_tree.bind("<<TreeviewSelect>>", self.on_category_select)

        product_box = ttk.LabelFrame(cat_product_pane, text="상품 리스트", padding=6)
        cat_product_pane.add(product_box, weight=4)

        product_top = ttk.Frame(product_box)
        product_top.pack(fill=tk.X, pady=(0, 6))
        ttk.Label(product_top, text="상품 검색").pack(side=tk.LEFT)
        self.product_search_var = tk.StringVar()
        self.product_search_var.trace_add("write", lambda *args: self.refresh_product_list())
        ttk.Entry(product_top, textvariable=self.product_search_var).pack(side=tk.LEFT, fill=tk.X, expand=True, padx=5)
        ttk.Button(product_top, text="링크 열기", command=self.open_selected_product_link).pack(side=tk.RIGHT, padx=(5, 0))
        ttk.Button(product_top, text="상품목록 저장", command=self.export_current_products).pack(side=tk.RIGHT, padx=(5, 0))
        ttk.Button(product_top, text="공급사 요약 저장", command=self.export_supplier_summary).pack(side=tk.RIGHT, padx=(5, 0))

        self.product_count_var = tk.StringVar(value="상품 0건")
        ttk.Label(product_box, textvariable=self.product_count_var, foreground="#555555").pack(fill=tk.X, pady=(0, 4))

        self.product_tree = ttk.Treeview(product_box, columns=tuple(PRODUCT_COLUMNS), show="headings", selectmode="browse")
        self._config_tree(self.product_tree, {
            "상품코드": 85,
            "상품명": 360,
            "카테고리": 360,
            "기본수량": 70,
            "공급가": 80,
            "판매가1": 80,
            "판매가7": 80,
            "마진율": 60,
            "진열": 60,
            "노출": 60,
            "최초등록일": 80,
            "최근수정일": 80,
            "상품링크": 280,
        })
        self.product_tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        prod_y = ttk.Scrollbar(product_box, orient=tk.VERTICAL, command=self.product_tree.yview)
        prod_x = ttk.Scrollbar(product_box, orient=tk.HORIZONTAL, command=self.product_tree.xview)
        self.product_tree.configure(yscrollcommand=prod_y.set, xscrollcommand=prod_x.set)
        prod_y.pack(side=tk.RIGHT, fill=tk.Y)
        prod_x.pack(side=tk.BOTTOM, fill=tk.X)
        self.product_tree.bind("<Double-1>", lambda event: self.open_selected_product_link())

    def _config_tree(self, tree: ttk.Treeview, widths: dict[str, int]):
        for col, width in widths.items():
            tree.heading(col, text=col, command=lambda c=col, t=tree: self.sort_treeview(t, c, False))
            tree.column(col, width=width, minwidth=40, anchor=tk.W if col not in {"상품수", "카테고리수", "개수"} else tk.E)

    # ---------------- Data loading ----------------
    def choose_folder_and_reload(self):
        selected = filedialog.askdirectory(title="엑셀 파일이 있는 폴더 선택")
        if selected:
            self.data_dir = Path(selected)
            self.path_label.configure(text=str(self.data_dir))
            self.load_data()

    def load_data(self):
        try:
            self.status_var.set("데이터 로딩 중...")
            self.update_idletasks()
            data_dir = self.data_dir
            if not (data_dir / REQUIRED_JOIN_FILE).exists():
                selected = filedialog.askdirectory(
                    title=f"{REQUIRED_JOIN_FILE} 파일이 있는 폴더를 선택하세요"
                )
                if not selected:
                    self.status_var.set("필수 파일을 찾지 못했습니다.")
                    return
                self.data_dir = Path(selected)
                self.path_label.configure(text=str(self.data_dir))
                data_dir = self.data_dir

            required_path = data_dir / REQUIRED_JOIN_FILE
            if not required_path.exists():
                messagebox.showerror("필수 파일 없음", f"{REQUIRED_JOIN_FILE} 파일을 찾을 수 없습니다.")
                return

            self.goods_df = read_excel_fast(required_path)
            self.goods_df = self._ensure_columns(self.goods_df, ["업체명", "업체코드", "상품명", "카테고리"])

            self.supplier_df = self._read_optional_excel(data_dir / OPTIONAL_SUPPLIER_FILE)
            self.major_df = self._read_optional_excel(data_dir / OPTIONAL_MAJOR_FILE)
            self.category_df = self._read_optional_excel(data_dir / OPTIONAL_CATEGORY_FILE)

            self._prepare_supplier_summary()
            self.refresh_supplier_list()
            self.clear_all_detail_views()
            self.status_var.set(
                f"로딩 완료: 상품 {len(self.goods_df):,}건 / 공급사 {len(self.supplier_summary_df):,}개"
            )
        except Exception as e:
            self.status_var.set("로딩 실패")
            messagebox.showerror("오류", f"데이터 로딩 중 오류가 발생했습니다.\n\n{e}\n\n{traceback.format_exc()}")

    def _read_optional_excel(self, path: Path) -> "pd.DataFrame":
        if path.exists():
            return read_excel_fast(path)
        return pd.DataFrame()

    def _ensure_columns(self, df: "pd.DataFrame", required_cols: list[str]) -> "pd.DataFrame":
        df = df.copy()
        for col in required_cols:
            if col not in df.columns:
                df[col] = ""
        return df

    def _prepare_supplier_summary(self):
        goods = self.goods_df.copy()
        for col in ["업체명", "업체코드", "카테고리"]:
            if col not in goods.columns:
                goods[col] = ""

        grouped = goods.groupby(["업체코드", "업체명"], dropna=False).agg(
            상품수=("상품명", "size"),
            카테고리수=("카테고리", lambda s: int(s.replace("", pd.NA).dropna().nunique())),
        ).reset_index()

        # 공급사 세부정보 파일에 상품수가 있으면 참고하되, 실제 상품수는 조인 파일 기준을 우선합니다.
        if not self.supplier_df.empty:
            info_cols = [c for c in ["업체명", "업체코드", "전화번호", "휴대폰번호", "이메일", "주소", "인쇄가능여부"] if c in self.supplier_df.columns]
            if "업체코드" in self.supplier_df.columns:
                info = self.supplier_df[info_cols].drop_duplicates(subset=["업체코드"], keep="first")
                grouped = grouped.merge(info, on="업체코드", how="left", suffixes=("", "_정보"))
                if "업체명_정보" in grouped.columns:
                    grouped["업체명"] = grouped["업체명"].mask(grouped["업체명"].eq(""), grouped["업체명_정보"])
            else:
                info = self.supplier_df[info_cols].drop_duplicates(subset=["업체명"], keep="first")
                grouped = grouped.merge(info, on="업체명", how="left", suffixes=("", "_정보"))

        grouped["상품수_sort"] = pd.to_numeric(grouped["상품수"], errors="coerce").fillna(0).astype(int)
        grouped["카테고리수_sort"] = pd.to_numeric(grouped["카테고리수"], errors="coerce").fillna(0).astype(int)
        grouped = grouped.sort_values(["상품수_sort", "업체명"], ascending=[False, True])
        self.supplier_summary_df = grouped

    # ---------------- Supplier list/detail ----------------
    def refresh_supplier_list(self):
        self._clear_tree(self.supplier_tree)
        if self.supplier_summary_df.empty:
            return
        keyword = self.supplier_search_var.get().strip().lower()
        df = self.supplier_summary_df.copy()
        if keyword:
            mask = (
                df.get("업체명", "").astype(str).str.lower().str.contains(keyword, na=False)
                | df.get("업체코드", "").astype(str).str.lower().str.contains(keyword, na=False)
            )
            df = df[mask]
        for _, row in df.iterrows():
            values = [
                clean_text(row.get("업체코드", "")),
                clean_text(row.get("업체명", "")),
                f"{int(row.get('상품수_sort', 0)):,}",
                f"{int(row.get('카테고리수_sort', 0)):,}",
            ]
            self.supplier_tree.insert("", tk.END, values=values)
        self.status_var.set(f"공급사 목록 {len(df):,}개 표시")

    def on_supplier_select(self, event=None):
        item_id = self._selected_item(self.supplier_tree)
        if not item_id:
            return
        values = self.supplier_tree.item(item_id, "values")
        self.selected_supplier_code = values[0]
        self.selected_supplier_name = values[1]
        self.current_category_filter = ""
        self.product_search_var.set("")
        self.category_filter_label.configure(text="필터 없음")
        self.show_supplier_detail()
        self.show_major_summary()
        self.show_category_summary()
        self.prepare_product_df()
        self.refresh_product_list()

    def show_supplier_detail(self):
        code = self.selected_supplier_code
        name = self.selected_supplier_name
        info = {}

        # 세부정보 파일 우선
        if not self.supplier_df.empty:
            sdf = self.supplier_df.copy()
            if "업체코드" in sdf.columns and code:
                hit = sdf[sdf["업체코드"].astype(str).eq(code)]
            else:
                hit = sdf[sdf.get("업체명", "").astype(str).eq(name)]
            if not hit.empty:
                r = hit.iloc[0].to_dict()
                for col in SUPPLIER_INFO_COLUMNS:
                    if col in r:
                        info[col] = clean_text(r.get(col, ""))

        # 없거나 부족하면 조인 파일에서 보완
        goods_hit = self.goods_df[self.goods_df.get("업체코드", "").astype(str).eq(code)] if code else pd.DataFrame()
        if goods_hit.empty and name:
            goods_hit = self.goods_df[self.goods_df.get("업체명", "").astype(str).eq(name)]
        if not goods_hit.empty:
            r = goods_hit.iloc[0].to_dict()
            join_cols = ["업체명", "업체코드", "상호명", "사업자번호", "대표자", "업태", "종목", "전화번호", "휴대폰번호", "팩스", "이메일", "우편번호", "주소"]
            for col in join_cols:
                if col in r and not info.get(col):
                    info[col] = clean_text(r.get(col, ""))
            info["상품수"] = f"{len(goods_hit):,}"
            info["카테고리수"] = f"{goods_hit.get('카테고리', pd.Series(dtype=str)).replace('', pd.NA).dropna().nunique():,}"

        lines = []
        for title in ["업체명", "업체코드", "상품수", "카테고리수", "상호명", "사업자번호", "대표자", "업태", "종목", "전화번호", "휴대폰번호", "팩스", "이메일", "우편번호", "주소", "인쇄가능여부", "한박스당수량", "한박스당배송비", "이미지사용", "상품등록일", "상품최근수정일", "발주링크"]:
            val = info.get(title, "")
            if val:
                lines.append(f"{title}: {val}")
        if not lines:
            lines.append("표시할 공급사 정보가 없습니다.")

        self.info_text.configure(state=tk.NORMAL)
        self.info_text.delete("1.0", tk.END)
        self.info_text.insert(tk.END, "\n".join(lines))
        self.info_text.configure(state=tk.DISABLED)

    # ---------------- Major/category/product views ----------------
    def show_major_summary(self):
        self._clear_tree(self.major_tree)
        name = self.selected_supplier_name
        code = self.selected_supplier_code

        rows = []
        # 카테고리 집계 파일이 있으면, 대분류별 상품수를 더 정확히 표시
        if not self.category_df.empty and "업체명" in self.category_df.columns:
            cdf = self.category_df[self.category_df["업체명"].astype(str).eq(name)].copy()
            if not cdf.empty:
                if "상품분류(대)" in cdf.columns and "개수" in cdf.columns:
                    cdf["개수_num"] = pd.to_numeric(cdf["개수"].str.replace(",", "", regex=False), errors="coerce").fillna(0).astype(int)
                    summary = cdf.groupby("상품분류(대)", dropna=False)["개수_num"].sum().reset_index()
                    summary = summary.sort_values("개수_num", ascending=False)
                    rows = [(clean_text(r["상품분류(대)"]), int(r["개수_num"])) for _, r in summary.iterrows()]

        if not rows:
            # 조인 파일에서 카테고리 첫 단계를 대분류로 계산
            gdf = self._supplier_goods_rows(code, name)
            if not gdf.empty and "카테고리" in gdf.columns:
                major = gdf["카테고리"].map(lambda x: clean_text(x).split(">")[0].strip() if clean_text(x) else "미분류")
                summary = major.value_counts().reset_index()
                summary.columns = ["대분류", "상품수"]
                rows = [(r["대분류"], int(r["상품수"])) for _, r in summary.iterrows()]

        for major, count in rows:
            self.major_tree.insert("", tk.END, values=(major or "미분류", f"{count:,}"))

    def show_category_summary(self):
        self._clear_tree(self.category_tree)
        name = self.selected_supplier_name
        code = self.selected_supplier_code
        rows = []

        if not self.category_df.empty and "업체명" in self.category_df.columns:
            cdf = self.category_df[self.category_df["업체명"].astype(str).eq(name)].copy()
            if not cdf.empty:
                if "개수" in cdf.columns:
                    cdf["개수_num"] = pd.to_numeric(cdf["개수"].str.replace(",", "", regex=False), errors="coerce").fillna(0).astype(int)
                    cdf = cdf.sort_values("개수_num", ascending=False)
                for _, r in cdf.iterrows():
                    rows.append((
                        clean_text(r.get("카테고리", "")),
                        clean_text(r.get("개수", "")),
                        clean_text(r.get("상품분류(대)", "")),
                        clean_text(r.get("상품분류(중)", "")),
                        clean_text(r.get("상품분류(소)", "")),
                        clean_text(r.get("상품분류(세)", "")),
                    ))

        if not rows:
            # 집계 파일이 없으면 조인 파일에서 즉시 계산
            gdf = self._supplier_goods_rows(code, name)
            if not gdf.empty and "카테고리" in gdf.columns:
                vc = gdf["카테고리"].map(clean_text).replace("", "미분류").value_counts().reset_index()
                vc.columns = ["카테고리", "개수"]
                for _, r in vc.iterrows():
                    parts = [p.strip() for p in clean_text(r["카테고리"]).split(">")]
                    parts += [""] * (4 - len(parts))
                    rows.append((r["카테고리"], f"{int(r['개수']):,}", parts[0], parts[1], parts[2], parts[3]))

        for row in rows:
            self.category_tree.insert("", tk.END, values=row)

    def prepare_product_df(self):
        self.current_product_df = self._supplier_goods_rows(self.selected_supplier_code, self.selected_supplier_name).copy()
        for col in PRODUCT_COLUMNS:
            if col not in self.current_product_df.columns:
                self.current_product_df[col] = ""
        self.current_product_df = self.current_product_df[PRODUCT_COLUMNS]

    def refresh_product_list(self):
        if self.current_product_df.empty:
            self._clear_tree(self.product_tree)
            self.product_count_var.set("상품 0건")
            return

        df = self.current_product_df.copy()
        keyword = self.product_search_var.get().strip().lower()
        if self.current_category_filter:
            df = df[df["카테고리"].astype(str).eq(self.current_category_filter)]
        if keyword:
            mask = (
                df.get("상품명", "").astype(str).str.lower().str.contains(keyword, na=False)
                | df.get("상품코드", "").astype(str).str.lower().str.contains(keyword, na=False)
                | df.get("카테고리", "").astype(str).str.lower().str.contains(keyword, na=False)
            )
            df = df[mask]

        self.current_product_filtered_df = df
        self._clear_tree(self.product_tree)
        # Treeview는 너무 많은 행을 한 번에 넣으면 느릴 수 있으므로 표시 상한을 둡니다.
        display_limit = 5000
        for _, row in df.head(display_limit).iterrows():
            values = [clean_text(row.get(col, "")) for col in PRODUCT_COLUMNS]
            self.product_tree.insert("", tk.END, values=values)
        suffix = "" if len(df) <= display_limit else f" / 화면 표시 {display_limit:,}건"
        self.product_count_var.set(f"상품 {len(df):,}건{suffix}")

    def on_category_select(self, event=None):
        item_id = self._selected_item(self.category_tree)
        if not item_id:
            return
        values = self.category_tree.item(item_id, "values")
        category = values[0]
        if not category:
            return
        self.current_category_filter = category
        self.category_filter_label.configure(text=f"필터: {category}")
        self.refresh_product_list()

    def on_major_select(self, event=None):
        item_id = self._selected_item(self.major_tree)
        if not item_id:
            return
        values = self.major_tree.item(item_id, "values")
        major = values[0]
        if not major:
            return
        # 대분류 클릭 시 상품 검색어를 대분류명으로 넣어 빠르게 필터링합니다.
        self.current_category_filter = ""
        self.category_filter_label.configure(text=f"대분류 검색: {major}")
        self.product_search_var.set(major)

    def clear_category_filter(self):
        self.current_category_filter = ""
        self.category_filter_label.configure(text="필터 없음")
        self.refresh_product_list()

    def _supplier_goods_rows(self, code: str, name: str) -> "pd.DataFrame":
        if self.goods_df.empty:
            return pd.DataFrame()
        df = self.goods_df
        if code and "업체코드" in df.columns:
            hit = df[df["업체코드"].astype(str).eq(code)]
            if not hit.empty:
                return hit
        if name and "업체명" in df.columns:
            return df[df["업체명"].astype(str).eq(name)]
        return pd.DataFrame()

    # ---------------- Actions ----------------
    def open_selected_product_link(self):
        item_id = self._selected_item(self.product_tree)
        if not item_id:
            messagebox.showinfo("안내", "상품을 선택하세요.")
            return
        values = self.product_tree.item(item_id, "values")
        col_index = PRODUCT_COLUMNS.index("상품링크")
        url = values[col_index]
        if not url:
            messagebox.showinfo("안내", "선택한 상품에 상품링크가 없습니다.")
            return
        webbrowser.open(url)

    def export_current_products(self):
        if self.current_product_filtered_df.empty:
            messagebox.showinfo("안내", "저장할 상품 목록이 없습니다.")
            return
        default_name = f"{self.selected_supplier_name}_상품목록_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
        default_name = self._safe_filename(default_name)
        path = filedialog.asksaveasfilename(
            title="상품목록 저장",
            defaultextension=".xlsx",
            initialfile=default_name,
            filetypes=[("Excel files", "*.xlsx")],
        )
        if not path:
            return
        try:
            self.current_product_filtered_df.to_excel(path, index=False)
            messagebox.showinfo("완료", f"저장 완료\n{path}")
        except Exception as e:
            messagebox.showerror("오류", f"저장 실패: {e}")

    def export_supplier_summary(self):
        if not self.selected_supplier_name:
            messagebox.showinfo("안내", "공급사를 선택하세요.")
            return
        default_name = f"{self.selected_supplier_name}_공급사요약_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
        default_name = self._safe_filename(default_name)
        path = filedialog.asksaveasfilename(
            title="공급사 요약 저장",
            defaultextension=".xlsx",
            initialfile=default_name,
            filetypes=[("Excel files", "*.xlsx")],
        )
        if not path:
            return
        try:
            info_df = self._supplier_info_as_df()
            major_df = self._tree_to_df(self.major_tree)
            category_df = self._tree_to_df(self.category_tree)
            products_df = self.current_product_filtered_df if not self.current_product_filtered_df.empty else self.current_product_df
            with pd.ExcelWriter(path, engine="openpyxl") as writer:
                info_df.to_excel(writer, index=False, sheet_name="공급사정보")
                major_df.to_excel(writer, index=False, sheet_name="대분류요약")
                category_df.to_excel(writer, index=False, sheet_name="카테고리요약")
                products_df.to_excel(writer, index=False, sheet_name="상품목록")
            messagebox.showinfo("완료", f"저장 완료\n{path}")
        except Exception as e:
            messagebox.showerror("오류", f"저장 실패: {e}")

    def _supplier_info_as_df(self) -> "pd.DataFrame":
        text = self.info_text.get("1.0", tk.END).strip()
        rows = []
        for line in text.splitlines():
            if ":" in line:
                k, v = line.split(":", 1)
                rows.append({"항목": k.strip(), "내용": v.strip()})
        return pd.DataFrame(rows)

    # ---------------- Utilities ----------------
    def _clear_tree(self, tree: ttk.Treeview):
        tree.delete(*tree.get_children())

    def _selected_item(self, tree: ttk.Treeview) -> str:
        selected = tree.selection()
        return selected[0] if selected else ""

    def _tree_to_df(self, tree: ttk.Treeview) -> "pd.DataFrame":
        cols = list(tree["columns"])
        rows = []
        for item in tree.get_children():
            values = list(tree.item(item, "values"))
            rows.append(dict(zip(cols, values)))
        return pd.DataFrame(rows)

    def _safe_filename(self, name: str) -> str:
        for ch in '\\/:*?"<>|':
            name = name.replace(ch, "_")
        return name

    def sort_treeview(self, tree: ttk.Treeview, col: str, reverse: bool):
        try:
            data = [(tree.set(k, col), k) for k in tree.get_children("")]
            def sort_key(item):
                value = item[0].replace(",", "")
                try:
                    return float(value)
                except Exception:
                    return value
            data.sort(key=sort_key, reverse=reverse)
            for index, (_, k) in enumerate(data):
                tree.move(k, "", index)
            tree.heading(col, command=lambda: self.sort_treeview(tree, col, not reverse))
        except Exception:
            pass

    def clear_all_detail_views(self):
        for tree in [self.major_tree, self.category_tree, self.product_tree]:
            self._clear_tree(tree)
        self.product_count_var.set("상품 0건")
        self.category_filter_label.configure(text="필터 없음")
        self.info_text.configure(state=tk.NORMAL)
        self.info_text.delete("1.0", tk.END)
        self.info_text.insert(tk.END, "왼쪽에서 공급사를 선택하세요.")
        self.info_text.configure(state=tk.DISABLED)


def main():
    app = SupplierProductViewer()
    app.mainloop()


if __name__ == "__main__":
    main()
