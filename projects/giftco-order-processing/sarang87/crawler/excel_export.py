# -*- coding: utf-8 -*-
"""
판촉사랑(87sarang) 크롤링 - 체크포인트 -> 엑셀 저장

크롤링(요청/파싱) 자체가 아니라 결과를 엑셀 파일로 내보내는 부분만 담당합니다.
"""
import logging
from copy import copy
from pathlib import Path

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from . import config
from .checkpoint_io import COLUMNS, load_checkpoint

logger = logging.getLogger(__name__)


def create_fallback_workbook():
    wb = Workbook()
    ws = wb.active
    ws.title = "납품사례_분류가격"

    ws.append(COLUMNS)

    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(name="맑은 고딕", size=10, bold=True, color="FFFFFF")
    thin_gray = Side(style="thin", color="D9E2F3")

    for col_idx, col_name in enumerate(COLUMNS, 1):
        cell = ws.cell(1, col_idx)
        cell.value = col_name
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = Border(
            left=thin_gray,
            right=thin_gray,
            top=thin_gray,
            bottom=thin_gray
        )

    widths = [5, 12, 18, 18, 18, 45, 20, 22, 22, 13, 13, 13, 13, 13, 15, 35, 13, 15, 45]
    for i, width in enumerate(widths, 1):
        ws.column_dimensions[ws.cell(1, i).column_letter].width = width

    ws.row_dimensions[1].height = 22
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{ws.cell(1, len(COLUMNS)).column_letter}1"

    return wb, ws


def save_excel_from_checkpoint(output_xlsx=None):
    if output_xlsx is None:
        output_xlsx = config.OUTPUT_XLSX

    checkpoint_df = load_checkpoint()

    if checkpoint_df.empty:
        logger.warning("[엑셀 저장 생략] 체크포인트 데이터가 없습니다.")
        return

    export_df = checkpoint_df.copy()

    for col in COLUMNS:
        if col not in export_df.columns:
            export_df[col] = ""

    export_df = export_df.reset_index(drop=True)
    export_df["No"] = range(1, len(export_df) + 1)
    export_df = export_df[COLUMNS]

    template_path = Path(config.TEMPLATE_XLSX)

    thin_gray = Side(style="thin", color="D9D9D9")
    body_border = Border(
        left=thin_gray,
        right=thin_gray,
        top=thin_gray,
        bottom=thin_gray
    )
    body_font = Font(name="맑은 고딕", size=10, color="000000")
    body_alignment = Alignment(vertical="center", wrap_text=False)
    no_fill = PatternFill(fill_type=None)

    if template_path.exists():
        wb = load_workbook(template_path)
        ws = wb.active

        # 기존 템플릿의 1행 헤더 스타일을 기준으로 보존
        # 새로 추가된 구매처분류 컬럼 등은 A1 스타일을 기준으로 적용
        header_styles = []
        header_fonts = []
        header_fills = []
        header_borders = []
        header_alignments = []

        for c in range(1, len(COLUMNS) + 1):
            source_col = c if c <= ws.max_column else 1
            header_styles.append(copy(ws.cell(1, source_col)._style))
            header_fonts.append(copy(ws.cell(1, source_col).font))
            header_fills.append(copy(ws.cell(1, source_col).fill))
            header_borders.append(copy(ws.cell(1, source_col).border))
            header_alignments.append(copy(ws.cell(1, source_col).alignment))

        if ws.max_row > 1:
            ws.delete_rows(2, ws.max_row - 1)

        for c, col_name in enumerate(COLUMNS, 1):
            cell = ws.cell(1, c)
            cell.value = col_name
            cell._style = copy(header_styles[c - 1])
            cell.font = copy(header_fonts[c - 1])
            cell.fill = copy(header_fills[c - 1])
            cell.border = copy(header_borders[c - 1])
            cell.alignment = copy(header_alignments[c - 1])

    else:
        wb, ws = create_fallback_workbook()

    for r_idx, record in enumerate(export_df.to_dict("records"), start=2):
        for c_idx, col_name in enumerate(COLUMNS, start=1):
            cell = ws.cell(r_idx, c_idx)
            value = record.get(col_name)

            if pd.isna(value):
                value = None

            # 가격 컬럼 처리
            # 가격문의는 문자열 그대로 유지하고, 숫자는 숫자로 저장
            if col_name in ["대량가격(원)", "중간가격(원)", "소량가격(원)"]:
                if value not in [None, ""]:
                    value_str = str(value).strip()

                    if "가격문의" in value_str:
                        value = "가격문의"
                    else:
                        value_str = value_str.replace(",", "").replace("원", "").strip()
                        value = int(value_str) if value_str.isdigit() else None

            cell.value = value

            # 2행부터는 배경색 제거
            cell.fill = no_fill
            cell.font = body_font
            cell.border = body_border
            cell.alignment = body_alignment

    price_cols = [COLUMNS.index(col) + 1 for col in ["대량가격(원)", "중간가격(원)", "소량가격(원)"]]
    text_cols = [
        COLUMNS.index(col) + 1
        for col in [
            "상품코드",
            "구매처분류(중)", "구매처분류(소)", "구매처분류(세)",
            "최소인쇄수량", "최소주문수량",
            "인쇄방법", "업종/행사", "등록일",
            "납품사례ID", "수집키"
        ]
        if col in COLUMNS
    ]

    # 숫자/텍스트 형식
    for row in range(2, ws.max_row + 1):
        for col in price_cols:
            if ws.cell(row, col).value == "가격문의":
                ws.cell(row, col).number_format = '@'
            else:
                ws.cell(row, col).number_format = '#,##0'

        for col in text_cols:
            ws.cell(row, col).number_format = '@'

    default_widths = [5, 12, 18, 18, 18, 45, 20, 22, 22, 13, 13, 13, 13, 13, 15, 35, 13, 15, 45]
    for i, width in enumerate(default_widths, 1):
        col_letter = ws.cell(1, i).column_letter
        current_width = ws.column_dimensions[col_letter].width

        if current_width is None or current_width < 5:
            ws.column_dimensions[col_letter].width = width

    ws.freeze_panes = "A2"
    last_col_letter = ws.cell(1, len(COLUMNS)).column_letter
    ws.auto_filter.ref = f"A1:{last_col_letter}{max(ws.max_row, 1)}"

    wb.save(output_xlsx)

    # 저장 후 output 폴더에도 동일 파일 복사
    config.sync_to_dir(output_xlsx)

    logger.info(f"[엑셀 저장 완료] {output_xlsx} / {len(export_df)}건")
    logger.debug(f"[output 폴더 복사 완료] {config.SAVE_DIR / Path(output_xlsx).name}")
