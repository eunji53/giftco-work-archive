import re
import datetime
from pathlib import Path
from openpyxl import load_workbook

# 01_collect_and_clean_mail.ipynb 실행 후 이 스크립트를 실행하세요.
# (01의 출력물인 01-10_mail_clean_final.xlsx 를 입력으로 사용합니다)

# ─────────────────────────────────────────────
# 헬퍼
# ─────────────────────────────────────────────
def extract_between(text, start_key, end_keys):
    idx = text.find(start_key)
    if idx < 0:
        return ""
    start = idx + len(start_key)
    end = len(text)
    for ek in end_keys:
        ei = text.find(ek, start)
        if ei > 0 and ei < end:
            end = ei
    return text[start:end].strip()


ORG_KEYWORDS = [
    "학교", "대학교", "대학", "병원", "의원", "센터", "공사", "재단", "협회", "조합",
    "주식회사", "기업", "연구소", "농협", "신협", "금고", "은행", "지점",
    "구청", "시청", "군청", "행정복지센터", "지원청", "교육청", "소방서",
    "경찰서", "복지관", "어린이집", "유치원", "코리아", "Korea", "법인", "협동조합",
    "관광", "공단", "사업단", "재단법인", "의회", "협의회", "연합회", "문화원",
    "도서관", "박물관", "미술관", "사무소", "출장소", "관리소",
    "주유소", "새마을금고", "출판", "미디어", "광고", "기획", "엔지니어링",
    "솔루션", "테크", "시스템", "네트워크", "정보", "물산", "상사",
    "서비스", "컨설팅", "연구원", "요양원", "요양센터", "요양병원",
    "주민센터", "복지센터", "상담센터", "지원센터", "커뮤니케이션", "커뮤니티",
    "인더스트리", "파트너스", "어소시에이츠", "그룹", "홀딩스",
]

# 직함 접미어 — 뒤에 이게 붙으면 개인 이름이 아니라 직함 포함 수령자
TITLE_SUFFIXES = [
    "님", "씨", "팀장", "과장", "대리", "사원", "부장", "이사", "사장", "대표",
    "원장", "소장", "관장", "교수", "선생", "강사", "연구원", "연구사",
    "주무관", "담당자", "담당", "책임", "수석", "실장", "본부장", "센터장",
    "지점장", "지사장", "국장", "처장", "차장", "매니저", "MD", "md",
]


def extract_buyer_and_contact(recipient):
    """
    수령자 문자열 → (구매처명, 담당자명) 분리

    - "전남대학교 / 김명숙"              → ("전남대학교", "김명숙")
    - "교사 곽승철(세종과학예술영재학교)"  → ("세종과학예술영재학교", "교사 곽승철")
    - "국립국제교육원(우은숙님)"           → ("국립국제교육원", "우은숙")
    - "소나테크(주) 이주은"               → ("소나테크(주)", "이주은")
    - "고려대 의료원 차유심"              → ("고려대 의료원", "차유심")
    - "이영진 센터장님"                   → ("", "이영진 센터장님")  — 개인+직함
    - "서울특별시립 남부노인전문요양원"    → ("서울특별시립 남부노인전문요양원", "")
    - "신현주"                           → ("", "신현주")
    """
    r = recipient.strip()
    if not r:
        return "", ""

    # 패턴1: "기관 / 이름"
    m = re.match(r"^(.+?)\s*/\s*(.+)$", r)
    if m:
        return m.group(1).strip(), m.group(2).strip()

    # 패턴2: "(주)기관 이름" — (주)로 시작하는 법인 + 뒤에 짧은 한글 이름
    m = re.match(r"^(\(주\).+?)\s+([가-힣]{2,4})$", r)
    if m:
        return m.group(1).strip(), m.group(2).strip()

    # 패턴3: "기관명(주) 이름"
    m = re.match(r"^(.+?\(주\))\s+([가-힣]{2,4})$", r)
    if m:
        return m.group(1).strip(), m.group(2).strip()

    # 패턴4: "XXX(YYY)" — 괄호가 완전히 닫힌 경우
    m = re.match(r"^(.+?)\((.+)\)\s*$", r)
    if m:
        left   = m.group(1).strip()
        inside = m.group(2).strip()

        if inside in ("주", "유한"):
            return r, ""

        is_person_inside = (
            inside.endswith("님") or
            inside.endswith("씨") or
            (len(inside) <= 4 and re.match(r"^[가-힣]+$", inside)) or
            re.match(r"^\d", inside)
        )
        if is_person_inside:
            return left, re.sub(r"[님씨]$", "", inside)
        else:
            return inside, left

    # 패턴5: 괄호가 안 닫힌 경우 "XXX(YYY" — 괄호 안 내용으로 기관 판단
    m = re.match(r"^(.+?)\(([^)]+)$", r)
    if m:
        left   = m.group(1).strip()
        inside = m.group(2).strip()
        # 괄호 안이 기관처럼 보이면
        has_org_kw = any(kw in inside for kw in ORG_KEYWORDS)
        if has_org_kw or len(inside) >= 4:
            return inside, left   # 기관(이름 형태)
        return r, ""

    # 패턴6: 기관 키워드 포함 — 마지막 토큰이 짧은 한글 이름이면 분리
    for kw in ORG_KEYWORDS:
        if kw in r:
            tokens = r.split()
            last = tokens[-1]
            if (
                len(tokens) >= 2
                and len(last) <= 4
                and re.match(r"^[가-힣]+$", last)
                and kw not in last
            ):
                return " ".join(tokens[:-1]), last
            return r, ""

    # 나머지 → 개인 이름 (직함 포함도 그대로 담당자로)
    return "", r


# ─────────────────────────────────────────────
# 메인 파서
# ─────────────────────────────────────────────
def normalize_text(text):
    """줄바꿈 제거 + 공백 정리 + 'T :' 같은 콜론 앞뒤 공백 표준화 (완성본 파서와 동일)"""
    text = text.replace("\n", " ").replace("\r", " ")
    text = re.sub(r"\s+", " ", text).strip()
    text = re.sub(r"메일\s*:", "메일 :", text)
    text = re.sub(r"T\s*:", "T:", text)
    text = re.sub(r"F\s*:", "F:", text)
    return text


# 라벨이 실제 본문에 등장하는 순서는 메일마다 조금씩 다를 수 있어서,
# 정규식 하나씩 걸기보다 "라벨 위치를 찾아 정렬 → 라벨 사이 텍스트를 값으로 슬라이싱"하는 방식이 훨씬 안정적임
LABELS_IN_ORDER = [
    "품 명", "옵 션", "수 량", "단 가", "금 액", "부 가 세", "합 계",
    "인쇄문구", "발 송 일", "결 제 일", "도착주소", "배송수단",
    "업체담당", "담 당 자", "입금계좌", "비 고", "첨부파일",
]

PRINT_METHOD_KEYWORDS = [
    "전사인쇄", "실크인쇄", "실크", "레이저인쇄", "레이저", "UV인쇄", "UV",
    "자수", "박인쇄", "불박", "에폭시", "패드인쇄", "칼라인쇄", "옵셋인쇄", "풀컬러",
]


def make_remark(option_text, note_text):
    """비고 컬럼에 들어갈 텍스트 = 옵션 + 실제 '비 고' 라벨 섹션"""
    parts = []
    if option_text:
        parts.append(f"옵션: {option_text}")
    if note_text:
        parts.append(f"발주서비고: {note_text}")
    return " / ".join(parts)


def extract_tel(section_text):
    """'T:, 010-1234-5678' 처럼 첫 슬롯이 비고 콤마 뒤에 번호가 오는 경우까지 대응"""
    m = re.search(r"T:([^,\s]*),?\s*([\d\-]+)?", section_text)
    if not m:
        return ""
    t1 = (m.group(1) or "").strip()
    t2 = (m.group(2) or "").strip()
    return t1 if t1 else t2


def parse_any_order(body, subject):
    """발주서 본문에서 필드를 추출. 발주서가 아니면 None 반환.

    라벨("품 명", "옵 션" 등)이 있는 정식 서식과, 라벨이 없는 자유형 서식 둘 다 처리한다.
    """
    if not body or body == "None":
        return None
    b = str(body)

    # 발주서 여부 판단 (마커 or 상품코드) — 이 부분은 기존과 동일
    marker_pos = -1
    for marker in [
        "발 주 서 ( 주문번호",
        "발 주 서 (주)명성",
        "발주서 (주)명성",
        "(주)명성 귀하",
    ]:
        p = b.find(marker)
        if p >= 0:
            marker_pos = p
            break

    code_pos = b.find("-183-")
    if marker_pos < 0 and code_pos < 0:
        return None

    work_body_raw = b[marker_pos:] if marker_pos >= 0 else b[max(0, code_pos - 200):]
    text = normalize_text(work_body_raw)

    # ── 수령자 원문 / 구매처·담당자 분리 ───────
    recipient_raw = ""
    for pat in [r"\(수령자:([^\)]+)\)", r"수령자[:\s]+([^\s\(]+)"]:
        m = re.search(pat, text)
        if m:
            recipient_raw = m.group(1).strip()
            break
    if not recipient_raw:
        m = re.search(r"\(수령자:([^\)]+)\)", subject or "")
        if m:
            recipient_raw = m.group(1).strip()
    buyer_name, contact_person = extract_buyer_and_contact(recipient_raw)

    # ── 주문번호 ──────────────────────────────
    m = re.search(r"주문번호\s*:\s*(\d+)", text)
    order_num = m.group(1) if m else ""

    # ── 주문일(메일 제목에서) ──────────────────
    m = re.search(r"\[\s*(\d{4})-(\d{2})-(\d{2})", subject or "")
    order_date = None
    if m:
        try:
            order_date = datetime.datetime(int(m.group(1)), int(m.group(2)), int(m.group(3)))
        except Exception:
            pass

    is_labeled = any(x in text for x in ["품 명", "옵 션", "수 량", "단 가", "금 액"])

    product = ""
    option = ""
    qty = unit_price = amount = total = vat = None
    extra_cost = 0
    send_date = None
    tel = ""
    manager = ""
    print_text = ""
    remark_raw = ""

    if is_labeled:
        # ── 라벨 위치를 전부 찾아 정렬 → 라벨 사이를 값으로 슬라이싱 ──
        positions = []
        for label in LABELS_IN_ORDER:
            idx = text.find(label)
            if idx != -1:
                positions.append((idx, label))
        positions.sort()

        sections = {}
        for i, (start, label) in enumerate(positions):
            end = positions[i + 1][0] if i + 1 < len(positions) else len(text)
            sections[label] = text[start + len(label):end].strip(" :")

        product = sections.get("품 명", "")
        option = sections.get("옵 션", "")

        qm = re.search(r"([\d,]+)\s*EA", sections.get("수 량", ""), re.IGNORECASE)
        qty = int(qm.group(1).replace(",", "")) if qm else None

        pm = re.search(r"[￦₩]?\s*([\d,]+)", sections.get("단 가", ""))
        unit_price = int(pm.group(1).replace(",", "")) if pm else None

        money_section = sections.get("금 액", "")
        am = re.search(r"[￦₩]?\s*([\d,]+)", money_section)
        amount = int(am.group(1).replace(",", "")) if am else None

        em = re.search(r"\(([^)]+)\)", money_section)
        if em:
            for n in re.findall(r"[\d,]+원", em.group(1)):
                extra_cost += int(n.replace(",", "").replace("원", ""))

        vm = re.search(r"[￦₩]?\s*([\d,]+)", sections.get("부 가 세", ""))
        vat = int(vm.group(1).replace(",", "")) if vm else None

        tm = re.search(r"[￦₩]?\s*([\d,]+)", sections.get("합 계", ""))
        total = int(tm.group(1).replace(",", "")) if tm else None

        dm = re.search(r"(\d{4})년\s*(\d{1,2})월\s*(\d{1,2})일", sections.get("발 송 일", ""))
        if dm:
            try:
                send_date = datetime.datetime(int(dm.group(1)), int(dm.group(2)), int(dm.group(3)))
            except Exception:
                send_date = None

        tel = extract_tel(sections.get("도착주소", ""))

        manager = sections.get("담 당 자", "")

        print_text = sections.get("인쇄문구", "")
        if print_text in ["없음", "-"]:
            print_text = ""

        remark_raw = sections.get("비 고", "")

    else:
        # ── 라벨이 없는 자유형 서식: 숫자/기호 등장 순서로 추정 ──
        qty_match = re.search(r"(\d[\d,]*)\s*EA", text, re.IGNORECASE)
        money_matches = list(re.finditer(
            r"[￦₩]\s*[\d,]+(?:\s*\([^)]*(?:배송비|추가비)[^)]*\))?", text
        ))
        date_matches = list(re.finditer(r"\d{4}년\s*\d{1,2}월\s*\d{1,2}일", text))
        option_match = re.search(r"\[([^\]]+)\]", text)
        receiver_match = re.search(r"([가-힣]{2,6})\s*T:\s*(01[0-9]-\d{3,4}-\d{4})", text)

        if qty_match:
            product = text[:qty_match.start()].strip()
            qty = int(qty_match.group(1).replace(",", ""))

        def money_val(match_obj):
            mm = re.search(r"([\d,]+)", match_obj.group(0))
            return int(mm.group(1).replace(",", "")) if mm else None

        if len(money_matches) >= 1:
            unit_price = money_val(money_matches[0])
        if len(money_matches) >= 2:
            amount = money_val(money_matches[1])
        if len(money_matches) >= 4:
            total = money_val(money_matches[3])
        elif len(money_matches) >= 3:
            total = money_val(money_matches[-1])

        if option_match:
            option = option_match.group(1).strip()

        if date_matches:
            dm = re.search(r"(\d{4})년\s*(\d{1,2})월\s*(\d{1,2})일", date_matches[0].group(0))
            try:
                send_date = datetime.datetime(int(dm.group(1)), int(dm.group(2)), int(dm.group(3)))
            except Exception:
                send_date = None

        if receiver_match:
            if not recipient_raw:
                recipient_raw = receiver_match.group(1)
                buyer_name, contact_person = extract_buyer_and_contact(recipient_raw)
            tel = receiver_match.group(2)
        if not tel:
            tel = extract_tel(text)

        note_m = re.search(r"(발송일이 .*?)(?:메일\s*:\s*[\w\.-]+@[\w\.-]+\.\w+)", text)
        remark_raw = note_m.group(1).strip() if note_m else ""

    # ── 인쇄방법 (라벨/자유형 공통, 본문 전체에서 탐색) ─────
    print_method = ""
    for kw in PRINT_METHOD_KEYWORDS:
        if kw in text:
            print_method = kw
            break

    # ── 선물포장 (붙여쓰기/띄어쓰기 둘 다 대응) ─────────
    gift_wrap = "O" if re.search(r"선물\s*포장", option or "") or re.search(r"선물\s*포장", text[:500]) else ""

    if not product and not qty:
        return None

    remark = make_remark(option, remark_raw)

    return {
        "order_num":      order_num,
        "order_date":     order_date,
        "send_date":      send_date,
        "recipient_raw":  recipient_raw,
        "buyer_name":     buyer_name,
        "contact_person": contact_person,
        "manager":        manager,
        "tel":            tel,
        "product":        product,
        "unit_price":     unit_price,
        "qty":            qty,
        "amount":         amount,
        "extra_cost":     extra_cost or None,
        "total":          total,
        "vat":            vat,
        "print_text":     print_text,
        "print_method":   print_method,
        "gift_wrap":      gift_wrap,
        "option":         option,
        "remark":         remark,
    }


# ─────────────────────────────────────────────
# 실행 — 경로 설정 (실행 전 확인 필요)
# ─────────────────────────────────────────────
BASE_DIR = Path(__file__).resolve().parent

# 01_collect_and_clean_mail.ipynb 의 결과물 (joagift/output/ 폴더에 저장됨)
INPUT_MAIL = BASE_DIR / "output" / "01-10_mail_clean_final.xlsx"

# 실제 거래데이터 템플릿 파일 경로 (02_parse_and_merge_test.ipynb와 동일하게 맞춤)
INPUT_TMPL = Path(r"C:\Users\USER\Desktop\김은지_업무파일\거래데이터\기프트코_거래데이터_수집양식.xlsx")

OUTPUT_PATH = BASE_DIR / "output" / "joagift_transaction_data.xlsx"


def get_mail_type(subject):
    """메일 제목에서 유형 판단: 원본 / Re(답장) / Fwd(전달)"""
    s = str(subject) if subject else ""
    if re.match(r"^(Re|RE)\s*:", s):
        return "Re(답장)"
    if re.match(r"^(FW|Fwd|FWD)\s*:", s, re.IGNORECASE):
        return "Fwd(전달)"
    if re.match(r"^(Re\s*:\s*)+(FW|Fwd)", s, re.IGNORECASE):
        return "Re(답장)"
    return "원본"


wb_mail = load_workbook(INPUT_MAIL, read_only=True)
rows = list(wb_mail.active.iter_rows(values_only=True))

parsed = []
for row in rows[1:]:
    result = parse_any_order(row[1], row[0])
    if result:
        result["mail_key"]  = row[2]
        result["uid"]       = row[3]
        result["mail_type"] = get_mail_type(row[0])
        parsed.append(result)

print(f"파싱 성공: {len(parsed)}건")

from collections import Counter
order_num_counter = Counter(d["order_num"] for d in parsed if d["order_num"])
for d in parsed:
    onum = d["order_num"]
    if onum and order_num_counter[onum] > 1:
        d["dup_flag"] = f"중복({order_num_counter[onum]}건)"
    else:
        d["dup_flag"] = ""

type_counts = Counter(d["mail_type"] for d in parsed)
dup_count   = sum(1 for d in parsed if d["dup_flag"])
print(f"  원본: {type_counts['원본']}건 / Re(답장): {type_counts['Re(답장)']}건 / Fwd(전달): {type_counts['Fwd(전달)']}건")
print(f"  주문번호 중복 포함 행: {dup_count}건")

wb_out = load_workbook(INPUT_TMPL)
ws_out = wb_out["거래데이터"]
for row_idx in range(ws_out.max_row, 4, -1):
    ws_out.delete_rows(row_idx)

ws_out.cell(1, 29, "메일유형")
ws_out.cell(1, 30, "주문번호중복")
ws_out.cell(1, 31, "메일키")
ws_out.cell(1, 32, "UID")
ws_out.cell(1, 33, "수령자원문")
ws_out.cell(1, 34, "조아기프트담당자(참고용)")

for i, d in enumerate(parsed, start=1):
    r = 4 + i
    vals = [
        i, d["order_num"], "조아기프트", d["order_date"], d["send_date"],
        "", "", d["buyer_name"], d["contact_person"], d["tel"],
        "", "", None, d["product"], "",
        "", "", "", d["unit_price"], "",
        d["qty"], d["amount"], d["extra_cost"], d["total"],
        d["print_method"], "", d["gift_wrap"], d["remark"],
        d["mail_type"], d["dup_flag"],
        d["mail_key"], d["uid"], d["recipient_raw"], d["manager"],
    ]
    for col, val in enumerate(vals, start=1):
        if val == "None" or val is None:
            val = None
        cell = ws_out.cell(row=r, column=col, value=val)
        if isinstance(val, datetime.datetime):
            cell.number_format = "YYYY-MM-DD"

OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
wb_out.save(OUTPUT_PATH)
print(f"저장 완료 → {OUTPUT_PATH}")
